#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
folha_fees.py — ingestão de Folha de pagamento e Fees por cliente (API_CONTRACT.md §Novas tabelas).

Fontes (por empresa, mesmas planilhas da DRE):
  * Abas 'Folha Jan'..'Folha Dez' — colunas DATA,NOME,Departamento,Cargo,TIPO,SALARIO,EXTRA,TOTAL
    (a linha de cabeçalho é DETECTADA — nem sempre é a primeira). Linhas em branco, blocos de
    resumo por departamento e a linha de total (NOME vazio/numérico) são ignorados.
  * Aba de fees — nome varia ('Fees', 'Fees ', 'Fees  '). Há dois blocos:
      - bloco esquerdo 'CLIENTE / CONTRATO' + valor mensal (cols B/C) — autoritativo SÓ na REF;
      - bloco direito 'CLIENTE' + 'VALOR TOTAL DE FEES' (ranking ABC) — nas demais empresas o
        bloco esquerdo é resíduo de template (BD/Viv/4PR trazem clientes de outra empresa e a
        Zup nem tem bloco esquerdo), então usamos o bloco DIREITO fora da REF.
    Linhas 'TOTAL BRUTO', notas de rodapé ('* ...') e linhas sem cliente são ignoradas.

Tabelas criadas via CREATE TABLE IF NOT EXISTS (exatamente como no API_CONTRACT.md):
  fato_folha_mensal, fato_fee_cliente, cockpit_alert_snooze.
Carga idempotente: DELETE por empresa+período (folha) / empresa+ano (fees) e INSERT.

Uso:  cd ingestao && python folha_fees.py
Env:  INCOMING (pasta das planilhas; default ../data/incoming)
"""
import os, sys, re, glob, pathlib

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from psycopg2.extras import execute_values

from db import get_conn, _empresa_id, _periodo_id

ROOT = pathlib.Path(__file__).resolve().parent.parent
INCOMING = pathlib.Path(os.environ.get("INCOMING", ROOT / "data" / "incoming"))

# empresa -> prefixo do arquivo (mesmo padrão de main.py / history.py)
PREFIX = {"REF": "REF+", "BD": "BD", "4PR": "4PR", "VIV": "Viv", "ZUP": "Zup"}
MESES = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

# DDL EXATAMENTE como no API_CONTRACT.md (§Novas tabelas)
DDL = """
CREATE TABLE IF NOT EXISTS fato_folha_mensal (
    id           SERIAL PRIMARY KEY,
    empresa_id   INT REFERENCES dim_empresa(id),
    periodo_id   INT REFERENCES dim_periodo(id),
    nome         VARCHAR(160),
    departamento VARCHAR(120),
    cargo        VARCHAR(120),
    tipo         VARCHAR(40),
    salario      NUMERIC(14,2),
    extra        NUMERIC(14,2),
    total        NUMERIC(14,2),
    total_mes    NUMERIC(14,2),
    UNIQUE (empresa_id, periodo_id, nome, departamento, cargo)
);
-- total_mes (col T "TOTAL MÊS" = custo total empresa). Migração idempotente
-- p/ DBs que já tinham a tabela só com 'total' (col H, bruto).
ALTER TABLE fato_folha_mensal ADD COLUMN IF NOT EXISTS total_mes NUMERIC(14,2);
CREATE TABLE IF NOT EXISTS fato_fee_cliente (
    id         SERIAL PRIMARY KEY,
    empresa_id INT,
    cliente    VARCHAR(160),
    fee_mensal NUMERIC(14,2),
    ano        INT,
    UNIQUE (empresa_id, cliente, ano)
);
CREATE TABLE IF NOT EXISTS cockpit_alert_snooze (
    alert_id VARCHAR(40) PRIMARY KEY,
    ate      DATE NOT NULL
);
"""


def _find_file(prefix):
    hits = sorted(glob.glob(str(INCOMING / f"{prefix}*DRE*.xlsx")))
    return hits[0] if hits else None


def _year_from_filename(path, default=2026):
    m = re.search(r"(20\d{2})", os.path.basename(path))
    return int(m.group(1)) if m else default


def _num(v):
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


# ---- Folha -------------------------------------------------------------------
def _folha_header(ws):
    """Detecta a linha de cabeçalho (contém NOME e TOTAL) e devolve (linha, {COLUNA: idx}).
    max_col vai até 30: a col T "TOTAL MÊS" (índice 19) fica além do bloco DATA..SEXO."""
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=10, max_col=30, values_only=True), 1):
        idx = {str(v).strip().upper(): i for i, v in enumerate(row) if isinstance(v, str) and v.strip()}
        if "NOME" in idx and "TOTAL" in idx:
            return r, idx
    return None, {}


def parse_folha(wb, path):
    """Lê as 12 abas 'Folha <Mês>' -> dicts com mes, nome, departamento, cargo, tipo,
    salario, extra, total (col H, bruto) e total_mes (col T, custo total empresa)."""
    year = _year_from_filename(path)
    sheets = {nm.strip().lower(): nm for nm in wb.sheetnames}
    out = []
    for mnum, mes in enumerate(MESES, 1):
        nm = sheets.get(f"folha {mes.lower()}")
        if not nm:
            continue
        ws = wb[nm]
        hrow, idx = _folha_header(ws)
        if not hrow:
            continue

        def g(row, col):
            i = idx.get(col)
            return row[i] if i is not None and i < len(row) else None

        for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
            nome, dep = g(row, "NOME"), g(row, "DEPARTAMENTO")
            total = _num(g(row, "TOTAL"))
            # linha válida = NOME e Departamento textuais + TOTAL numérico.
            # Exclui: linhas em branco, linha de total (NOME vazio ou numérico) e o bloco de
            # resumo por departamento no rodapé (texto na col A, número na col B=NOME).
            if not isinstance(nome, str) or not nome.strip() or not isinstance(dep, str) or not dep.strip():
                continue
            if total is None or nome.strip().upper().startswith("TOTAL"):
                continue
            cargo, tipo = g(row, "CARGO"), g(row, "TIPO")
            # col T "TOTAL MÊS" = CUSTO TOTAL p/ a empresa (bruto+VT+VR+FGTS+INSS);
            # aceita grafia com/sem acento; fallback = total (bruto) se a coluna faltar.
            total_mes = _num(g(row, "TOTAL MÊS"))
            if total_mes is None:
                total_mes = _num(g(row, "TOTAL MES"))
            if total_mes is None:
                total_mes = total
            out.append({
                "year": year, "month": mnum,
                "nome": nome.strip()[:160], "departamento": dep.strip()[:120],
                "cargo": (str(cargo).strip()[:120] if cargo is not None else None),
                "tipo": (str(tipo).strip()[:40] if tipo is not None else None),
                "salario": _num(g(row, "SALARIO")) or 0.0,
                "extra": _num(g(row, "EXTRA")) or 0.0,
                "total": total,
                "total_mes": total_mes,
            })
    return out


# ---- Fees --------------------------------------------------------------------
def _clean_cliente(s):
    return re.sub(r"\*+$", "", s.strip()).strip()[:160]


def parse_fees(wb, empresa):
    """Lê a aba de fees -> [(cliente, fee_mensal)].
    REF: bloco esquerdo ('CLIENTE / CONTRATO', valor na coluna seguinte).
    Demais: bloco direito ('CLIENTE' + 'VALOR TOTAL DE FEES'), pois o esquerdo é template."""
    fee_sheet = next((nm for nm in wb.sheetnames if nm.strip().lower() == "fees"), None)
    if not fee_sheet:
        return []
    ws = wb[fee_sheet]

    left_c = right_c = header_r = None
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=10, max_col=20, values_only=True), 1):
        for i, v in enumerate(row):
            if not isinstance(v, str):
                continue
            t = v.strip().upper()
            if t == "CLIENTE / CONTRATO":
                left_c, header_r = i, r
            elif t == "VALOR TOTAL DE FEES":
                right_c, header_r = i, r          # coluna do cliente = right_c - 1
        if header_r:
            break
    if header_r is None:
        return []

    use_left = (empresa == "REF" and left_c is not None) or right_c is None
    cli_col = left_c if use_left else right_c - 1
    val_col = (left_c + 1) if use_left else right_c

    fees = []
    for row in ws.iter_rows(min_row=header_r + 1, values_only=True):
        cli = row[cli_col] if cli_col < len(row) else None
        val = _num(row[val_col]) if val_col < len(row) else None
        if not isinstance(cli, str) or not cli.strip() or val is None:
            continue
        t = cli.strip().upper()
        if t.startswith("TOTAL") or cli.strip().startswith("*"):     # totais e notas de rodapé
            continue
        fees.append((_clean_cliente(cli), val))
    return fees


# ---- carga -------------------------------------------------------------------
def upsert_folha(cur, empresa, rows, ec, pc):
    """DELETE por empresa+período e INSERT (idempotente). Duplicatas de chave são somadas."""
    if not rows:
        return 0
    emp_id = _empresa_id(cur, empresa, ec)
    periodos = sorted({(r["year"], r["month"]) for r in rows})
    pids = [_periodo_id(cur, y, m, pc) for y, m in periodos]
    cur.execute("DELETE FROM fato_folha_mensal WHERE empresa_id=%s AND periodo_id = ANY(%s)",
                (emp_id, pids))
    agg = {}
    for r in rows:
        pid = _periodo_id(cur, r["year"], r["month"], pc)
        k = (emp_id, pid, r["nome"], r["departamento"], r["cargo"])
        if k in agg:                                    # mesma pessoa 2x na aba -> soma
            a = agg[k]
            agg[k] = (a[0], r["tipo"], a[2] + r["salario"], a[3] + r["extra"],
                      a[4] + r["total"], a[5] + r["total_mes"])
        else:
            agg[k] = (k, r["tipo"], r["salario"], r["extra"], r["total"], r["total_mes"])
    valores = [(k[0], k[1], k[2], k[3], k[4], v[1], round(v[2], 2), round(v[3], 2),
                round(v[4], 2), round(v[5], 2))
               for k, v in agg.items()]
    execute_values(cur, """
        INSERT INTO fato_folha_mensal (empresa_id, periodo_id, nome, departamento, cargo,
                                       tipo, salario, extra, total, total_mes) VALUES %s
    """, valores)
    return len(valores)


def upsert_fees(cur, empresa, fees, ano, ec):
    """DELETE por empresa+ano e INSERT (idempotente). Cliente duplicado -> soma."""
    emp_id = _empresa_id(cur, empresa, ec)
    if not fees:
        return 0
    cur.execute("DELETE FROM fato_fee_cliente WHERE empresa_id=%s AND ano=%s", (emp_id, ano))
    agg = {}
    for cli, val in fees:
        agg[cli] = agg.get(cli, 0.0) + val
    execute_values(cur, "INSERT INTO fato_fee_cliente (empresa_id, cliente, fee_mensal, ano) VALUES %s",
                   [(emp_id, cli, round(v, 2), ano) for cli, v in agg.items()])
    return len(agg)


def run():
    """Executa a ingestão de folha+fees IN-PROCESS e devolve {ok, output} (padrão de main.run())."""
    out = [f"== Folha + Fees == (planilhas em {INCOMING})"]
    ok = True
    con = get_conn(); con.autocommit = False
    try:
        cur = con.cursor()
        cur.execute(DDL)                      # tabelas do API_CONTRACT.md (idempotente)
        con.commit()
        ec, pc = {}, {}
        for empresa, prefix in PREFIX.items():
            caminho = _find_file(prefix)
            if not caminho:
                out.append(f"  [{empresa}] arquivo não encontrado (prefixo {prefix})")
                ok = False
                continue
            try:
                wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
                try:
                    folha = parse_folha(wb, caminho)
                    fees = parse_fees(wb, empresa)
                finally:
                    wb.close()
                nf = upsert_folha(cur, empresa, folha, ec, pc)
                nc = upsert_fees(cur, empresa, fees, _year_from_filename(caminho), ec)
                con.commit()
                meses = len({(r["year"], r["month"]) for r in folha})
                out.append(f"  [{empresa}] OK: folha {nf} linhas ({meses} meses) | fees {nc} clientes")
                if nf == 0 or nc == 0:
                    out.append(f"  [{empresa}] ALERTA: folha ou fees vazios")
                    ok = False
            except Exception as e:
                con.rollback()
                out.append(f"  [{empresa}] ERRO: {e}")
                ok = False
    finally:
        con.close()
    return {"ok": ok, "output": "\n".join(out)}


def main():
    r = run()
    print(r["output"])
    sys.exit(0 if r["ok"] else 1)


if __name__ == "__main__":
    main()
