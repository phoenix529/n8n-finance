#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
history.py — carga do histórico anual 2018–2025 (Technical Blueprint §6.5 "todos os anos").

O DRE mensal detalhado só existe para 2026; os anos anteriores estão em abas de
resumo ('Resumo 18 a 26', 'Resumo 21 a 26', ...) como TOTAIS ANUAIS por linha.
Carregamos cada total anual como um ponto em dezembro daquele ano (mes=12) no
fato_dre_mensal — sem colidir com 2026 (que é mensal). Painel de evolução agrega por ano.

Uso:  cd ingestao && python history.py
"""
import os, sys, glob, pathlib, re
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from parsers.base import norm, CANON
from db import upsert_dre

ROOT = pathlib.Path(__file__).resolve().parent.parent
INCOMING = pathlib.Path(os.environ.get("INCOMING", ROOT / "data" / "incoming"))
PREFIX = {"REF": "REF+", "BD": "BD", "4PR": "4PR", "VIV": "Viv", "ZUP": "Zup"}


def _find_total_sheet(wb):
    """Acha a aba de resumo anual: linha de cabeçalho com >=2 células 'TOTAL 20xx'."""
    for nm in wb.sheetnames:
        ws = wb[nm]
        for r, row in enumerate(ws.iter_rows(min_row=1, max_row=6, max_col=14, values_only=True), 1):
            years = {}
            for ci, v in enumerate(row, 1):
                if isinstance(v, str):
                    m = re.search(r"TOTAL\s*(20\d{2})", v.upper())
                    if m:
                        years[ci] = int(m.group(1))
            if len(years) >= 2:
                return ws, r, years
    return None, None, {}


def load_company_history(empresa, path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws, hrow, years = _find_total_sheet(wb)
    if not ws:
        wb.close(); return []
    rows, taken = [], set()
    for row in ws.iter_rows(min_row=hrow + 1, max_row=ws.max_row, values_only=True):
        label = row[0] if row else None          # coluna A nas abas de resumo
        if label is None or not str(label).strip():
            continue
        for descricao, grupo, tipo, pats in CANON:
            if descricao in taken:
                continue
            # casa pelo mesmo conjunto de padrões da DRE-Base
            n = norm(label)
            if any((mode == "eq" and n == p) or (mode == "pre" and n.startswith(p)) for p, mode in pats):
                taken.add(descricao)
                for ci, yr in years.items():
                    if yr >= 2026:           # 2026 já é mensal
                        continue
                    v = row[ci - 1] if ci - 1 < len(row) else None
                    if isinstance(v, (int, float)):
                        rows.append({"company": empresa, "year": yr, "month": 12,
                                     "account_code": None, "account_description": descricao,
                                     "group": grupo, "value": round(float(v), 4), "source": os.path.basename(path)})
                break
    wb.close()
    return rows


def run():
    """Executa a carga do histórico IN-PROCESS e devolve {ok, output} (padrão de main.run())."""
    import pandas as pd
    out, ok, total = [], True, 0
    for emp, prefix in PREFIX.items():
        hits = sorted(glob.glob(str(INCOMING / f"{prefix}*DRE*.xlsx")))
        if not hits:
            out.append(f"  [{emp}] arquivo não encontrado (prefixo {prefix})")
            ok = False
            continue
        try:
            rows = load_company_history(emp, hits[0])
        except Exception as e:
            out.append(f"  [{emp}] ERRO: {e}")
            ok = False
            continue
        if rows:
            df = pd.DataFrame(rows)
            n = upsert_dre(emp, df)
            total += n
            anos = sorted({r["year"] for r in rows})
            out.append(f"  [{emp}] histórico: {n} linhas | anos {anos[0]}–{anos[-1]}")
        else:
            out.append(f"  [{emp}] sem aba de resumo anual detectada")
    out.append(f"-- histórico carregado: {total} linhas (2018–2025, ponto em dez) --")
    return {"ok": ok, "output": "\n".join(out)}


def main():
    r = run()
    print(r["output"])
    sys.exit(0 if r["ok"] else 1)


if __name__ == "__main__":
    main()
