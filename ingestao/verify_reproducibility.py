#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
verify_reproducibility.py — checa o critério de aceite §6.5 (reprodutibilidade):
o total no banco bate com o "TOTAL 2026" da própria planilha (tolerância R$ 0,01),
para AS 5 EMPRESAS e as linhas-chave da DRE.
"""
import os, sys, glob
import openpyxl
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from parsers.base import norm, CANON
from history import _find_total_sheet
from db import get_conn

INCOMING = os.path.join(os.path.dirname(__file__), "..", "data", "incoming")
PREFIX = {"REF": "REF+", "BD": "BD", "4PR": "4PR", "VIV": "Viv", "ZUP": "Zup"}
CHECK = ["RECEITA BRUTA", "RECEITA OPERACIONAL LIQUIDA",
         "RESULTADO OPERACIONAL ANTES DOS IMPOSTOS (EBIT)", "RESULTADO LIQUIDO"]
# padrões por descrição canônica (do CANON)
PATS = {d: pats for d, g, t, pats in CANON}


def planilha_total_2026(path):
    """Extrai os TOTAIS 2026 das linhas-chave da aba de resumo anual."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws, hrow, years = _find_total_sheet(wb)
    out = {}
    if ws:
        col2026 = next((ci for ci, y in years.items() if y == 2026), None)
        if col2026:
            taken = set()
            for row in ws.iter_rows(min_row=hrow + 1, max_row=ws.max_row, values_only=True):
                lab = row[0] if row else None
                if not lab or not str(lab).strip():
                    continue
                n = norm(lab)
                for desc in CHECK:
                    if desc in taken:
                        continue
                    if any((m == "eq" and n == p) or (m == "pre" and n.startswith(p)) for p, m in PATS[desc]):
                        v = row[col2026 - 1] if col2026 - 1 < len(row) else None
                        if isinstance(v, (int, float)):
                            out[desc] = float(v)
                        taken.add(desc)
                        break
    wb.close()
    return out


def db_total_2026(emp, desc):
    con = get_conn(); cur = con.cursor()
    cur.execute("""SELECT COALESCE(SUM(f.valor),0) FROM fato_dre_mensal f
                   JOIN dim_conta c ON c.id=f.conta_id JOIN dim_empresa e ON e.id=f.empresa_id
                   JOIN dim_periodo p ON p.id=f.periodo_id
                   WHERE e.codigo=%s AND c.descricao=%s AND p.ano=2026""", (emp, desc))
    v = float(cur.fetchone()[0]); cur.close(); con.close()
    return v


def main():
    print("=" * 92)
    print("REPRODUTIBILIDADE — DB (soma 2026) vs planilha (TOTAL 2026) — tolerância R$ 0,01")
    print("=" * 92)
    all_ok = True
    for emp, prefix in PREFIX.items():
        hits = sorted(glob.glob(os.path.join(INCOMING, f"{prefix}*DRE*.xlsx")))
        if not hits:
            print(f"{emp}: arquivo não encontrado"); all_ok = False; continue
        plan = planilha_total_2026(hits[0])
        print(f"\n■ {emp}")
        for desc in CHECK:
            pv = plan.get(desc)
            dv = db_total_2026(emp, desc)
            if pv is None:
                print(f"   {desc[:46]:46}  planilha=n/d   db={dv:>16,.2f}")
                continue
            diff = abs(dv - pv)
            ok = diff <= 0.01
            all_ok = all_ok and ok
            flag = "OK " if ok else "XX "
            print(f"   {flag}{desc[:43]:43} planilha={pv:>15,.2f}  db={dv:>15,.2f}  Δ={diff:,.2f}")
    print("\n" + ("VEREDITO: REPRODUTIBILIDADE OK (todas as linhas dentro de R$ 0,01)"
                  if all_ok else "VEREDITO: HÁ DIVERGÊNCIAS — ver linhas marcadas com XX"))
    sys.exit(0 if all_ok else 1)


if __name__ == "__main__":
    main()
