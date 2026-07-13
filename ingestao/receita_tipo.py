#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
receita_tipo.py — ingestão do MIX de RECEITA BRUTA por tipo (Painel 02).

Fonte: aba 'DRE-Base' de cada planilha em data/incoming/<Empresa>*DRE*.xlsx.
Estrutura (verificada): a linha 'RECEITA BRUTA' (col B) traz a RECEITA BRUTA do mês
nas colunas de mês (Jan..Dez); LOGO ABAIXO vêm as linhas de TIPO de receita (col B),
cuja SOMA por mês == RECEITA BRUTA do mês. Ex. REF+ 2026/Jan:
  FEE MENSAL + MIDIA OFF + MIDIA ONLINE & PROGRAMATICA + PRODUÇÃO GRAFICA + CRIAÇÃO
  + FILMES/SPOT + BVS + OUTROS RECEBIVEIS = 20.469.482,58 = RECEITA BRUTA.
Cada empresa tem um subconjunto (4PR só FEE+PRODUÇÃO+MIDIA; Viv sem FILMES/BVS; Zup sem FEE).

A coleta das linhas de tipo começa na linha seguinte à 'RECEITA BRUTA' e PARA ao encontrar
um rótulo com 'LIQUIDA'/'DEDU'/'OPERACIONAL'/'RESULTADO' ou uma linha vazia (fim do bloco).

Categorias CANÔNICAS (ordem fixa) — regra de rollup DEFAULT do Grupo (a confirmar c/ cliente):
  'Fee Mensal'  <= rótulo contém 'FEE'
  'Mídia Off'   <= 'MIDIA OFF'  ou ('MIDIA' sem ON/ONLINE/PROGRAMA)
  'Mídia On'    <= 'ONLINE' / 'PROGRAMA' / 'MIDIA ON'
  'Criação'     <= 'CRIA'
  'Filmes/Spot' <= 'FILME' / 'SPOT'
  'BVS'         <= 'BVS'
  'Outras'      <= qualquer outro (ex.: 'PRODUÇÃO GRAFICA', 'OUTROS RECEBIVEIS')

Destino: fato_receita_tipo_mensal(empresa_id, periodo_id, tipo, valor) — CREATE TABLE IF NOT
EXISTS + upsert idempotente (ON CONFLICT DO UPDATE). Multi-ano: um workbook por ano (ano do nome).

RECONCILE: após a carga, valida por empresa/ano que SUM(tipos) por mês == RECEITA BRUTA em
fato_dre_mensal (tolerância R$ 1). Divergências são reportadas (warn) — não derrubam a carga.

Uso:  cd ingestao && python receita_tipo.py
Env:  INCOMING (pasta das planilhas; default ../data/incoming)
"""
import os, sys, re, glob, pathlib

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from psycopg2.extras import execute_values

from db import get_conn, _empresa_id, _periodo_id
from parsers.base import _find_dre_sheet, _find_header, norm

ROOT = pathlib.Path(__file__).resolve().parent.parent
INCOMING = pathlib.Path(os.environ.get("INCOMING", ROOT / "data" / "incoming"))

# empresa -> prefixo do arquivo (mesmo padrão de main.py / folha_fees.py / tri_hist.py)
PREFIX = {"REF": "REF+", "BD": "BD", "4PR": "4PR", "VIV": "Viv", "ZUP": "Zup"}

# ordem fixa das categorias canônicas (contrato do Painel 02)
CATEGORIAS = ["Fee Mensal", "Mídia Off", "Mídia On", "Criação", "Filmes/Spot", "BVS", "Outras"]

# fim do bloco de tipos de receita (rótulos que já pertencem à DRE abaixo)
_STOP = ("liquida", "dedu", "operacional", "resultado")

DDL = """
CREATE TABLE IF NOT EXISTS fato_receita_tipo_mensal (
    id         SERIAL PRIMARY KEY,
    empresa_id INT NOT NULL REFERENCES dim_empresa(id),
    periodo_id INT NOT NULL REFERENCES dim_periodo(id),
    tipo       VARCHAR(40) NOT NULL,
    valor      NUMERIC(16,2),
    UNIQUE (empresa_id, periodo_id, tipo)
);
CREATE INDEX IF NOT EXISTS ix_rec_tipo_emp_per ON fato_receita_tipo_mensal(empresa_id, periodo_id);
"""


def _find_file(prefix):
    return sorted(glob.glob(str(INCOMING / f"{prefix}*DRE*.xlsx")))


def _year_from_filename(path, default=2026):
    m = re.search(r"(20\d{2})", os.path.basename(path))
    return int(m.group(1)) if m else default


def _num(v):
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def _categoria(label):
    """Mapeia o rótulo do tipo de receita -> categoria canônica (regra de rollup DEFAULT)."""
    n = norm(label)                                    # minúsculas, sem acento, espaços colapsados
    if "fee" in n:
        return "Fee Mensal"
    if "online" in n or "programa" in n or "midia on" in n:
        return "Mídia On"
    if "midia off" in n or "midia" in n:
        return "Mídia Off"
    if "cria" in n:
        return "Criação"
    if "filme" in n or "spot" in n:
        return "Filmes/Spot"
    if "bvs" in n:
        return "BVS"
    return "Outras"


def _is_receita_bruta(label):
    n = norm(label)
    return n == "receita" or n.startswith("receita bruta")


def parse_receita_tipo(path, year):
    """Lê a aba DRE-Base e devolve {mes: {categoria: valor}} (soma dos tipos por mês)."""
    src = os.path.basename(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = _find_dre_sheet(wb)
        if ws is None:
            raise RuntimeError(f"{src}: aba DRE-Base não encontrada")
        hrow, months = _find_header(ws, year)
        if not months:
            raise RuntimeError(f"{src}: cabeçalho de meses não encontrado")

        # 1) localizar a linha 'RECEITA BRUTA'
        rb_row = None
        for r, row in enumerate(ws.iter_rows(min_row=hrow + 1, max_row=ws.max_row, values_only=True), hrow + 1):
            label = row[1] if len(row) > 1 else None    # coluna B
            if isinstance(label, str) and _is_receita_bruta(label):
                rb_row = r
                break
        if rb_row is None:
            raise RuntimeError(f"{src}: linha 'RECEITA BRUTA' não encontrada")

        # 2) coletar as linhas de TIPO logo abaixo, até stop-rule / linha vazia
        mix = {m: {} for m in months.values()}          # {mes: {categoria: soma}}
        for row in ws.iter_rows(min_row=rb_row + 1, max_row=ws.max_row, values_only=True):
            label = row[1] if len(row) > 1 else None      # coluna B
            if label is None or not str(label).strip():   # linha vazia -> fim do bloco
                break
            n = norm(label)
            if any(tok in n for tok in _STOP):             # já é a DRE abaixo -> fim do bloco
                break
            cat = _categoria(label)
            for ci, mes in months.items():
                v = _num(row[ci - 1]) if ci - 1 < len(row) else None
                if v is None:
                    continue
                mix[mes][cat] = mix[mes].get(cat, 0.0) + v
        return mix
    finally:
        wb.close()


def upsert_receita_tipo(cur, empresa, mix, year, ec, pc):
    """Upsert idempotente em fato_receita_tipo_mensal. Retorna nº de linhas."""
    emp_id = _empresa_id(cur, empresa, ec)
    valores = []
    for mes, cats in mix.items():
        pid = _periodo_id(cur, year, mes, pc)
        for cat, val in cats.items():
            valores.append((emp_id, pid, cat, round(val, 2)))
    if not valores:
        return 0
    execute_values(cur, """
        INSERT INTO fato_receita_tipo_mensal (empresa_id, periodo_id, tipo, valor) VALUES %s
        ON CONFLICT (empresa_id, periodo_id, tipo) DO UPDATE SET valor = EXCLUDED.valor
    """, valores)
    return len(valores)


def _reconcile(cur, empresa, mix, year, ec):
    """Compara SUM(tipos) por mês (mix) com RECEITA BRUTA em fato_dre_mensal (tolerância R$ 1).
    Devolve lista de strings com as divergências (vazia = tudo bate)."""
    emp_id = _empresa_id(cur, empresa, ec)
    cur.execute("""
        SELECT p.mes, f.valor
        FROM fato_dre_mensal f
        JOIN dim_conta c   ON c.id = f.conta_id
        JOIN dim_periodo p ON p.id = f.periodo_id
        WHERE f.empresa_id = %s AND p.ano = %s AND c.descricao = 'RECEITA BRUTA'
    """, (emp_id, year))
    bruta = {int(m): float(v or 0.0) for m, v in cur.fetchall()}
    difs = []
    for mes, cats in sorted(mix.items()):
        soma = round(sum(cats.values()), 2)
        rb = bruta.get(mes)
        if rb is None:
            if abs(soma) > 1.0:
                difs.append(f"mês {mes}: RECEITA BRUTA ausente em fato_dre_mensal (tipos={soma:.2f})")
            continue
        if abs(soma - rb) > 1.0:
            difs.append(f"mês {mes}: tipos={soma:.2f} != bruta={rb:.2f} (Δ={soma - rb:.2f})")
    return difs


def run():
    """Executa a ingestão IN-PROCESS e devolve {ok, output} (padrão de folha_fees.run())."""
    out = [f"== Receita por tipo (fato_receita_tipo_mensal) == (planilhas em {INCOMING})"]
    ok = True
    con = get_conn(); con.autocommit = False
    try:
        cur = con.cursor()
        cur.execute(DDL)                      # tabela do Painel 02 (idempotente)
        con.commit()
        ec, pc = {}, {}
        for empresa, prefix in PREFIX.items():
            caminhos = _find_file(prefix)
            if not caminhos:
                out.append(f"  [{empresa}] arquivo não encontrado (prefixo {prefix})")
                ok = False
                continue
            for caminho in caminhos:                    # multi-ano: um workbook por ano
                ano = _year_from_filename(caminho)
                try:
                    mix = parse_receita_tipo(caminho, ano)
                    n = upsert_receita_tipo(cur, empresa, mix, ano, ec, pc)
                    con.commit()
                    difs = _reconcile(cur, empresa, mix, ano, ec)
                    if n == 0:
                        out.append(f"  [{empresa}/{ano}] ALERTA: nenhuma linha de tipo carregada")
                        ok = False
                    elif difs:
                        out.append(f"  [{empresa}/{ano}] PARCIAL: {n} linhas | reconc. FALHOU: {'; '.join(difs)}")
                        ok = False
                    else:
                        out.append(f"  [{empresa}/{ano}] OK: {n} linhas | reconc. OK (soma tipos == RECEITA BRUTA)")
                except Exception as e:
                    con.rollback()
                    out.append(f"  [{empresa}/{ano}] ERRO: {e}")
                    ok = False
    finally:
        con.close()
    return {"ok": ok, "output": "\n".join(out)}


def main():
    r = run()
    print(r["output"])
    sys.exit(0 if r["ok"] else 1)


if __name__ == "__main__":
    main()
