# -*- coding: utf-8 -*-
"""
parsers/clientes.py — receita por cliente da REF (Technical Blueprint §5.2, §7.3).

Fonte: aba 'Resumo Faturamento' (cliente × tipo de receita: FEE e VARIÁVEL),
que é o resumo limpo de faturamento por cliente de 2026. Devolve um DataFrame:
  company, year, month, cliente, tipo_receita, value, source
O período é gravado como 2026-12 (acumulado do ano), pois a aba traz totais 2026.
"""
import os
import pandas as pd
import openpyxl
from .base import norm

REF_YEAR = 2026
REF_MONTH = 12   # marcador "acumulado 2026"


def parse_clientes_ref(path, year=REF_YEAR):
    src = os.path.basename(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "Resumo Faturamento" not in wb.sheetnames:
        wb.close()
        return pd.DataFrame(columns=["company", "year", "month", "cliente", "tipo_receita", "value", "source"])
    ws = wb["Resumo Faturamento"]
    # localizar cabeçalho (linha com CLIENTES + FEES/VARIAVEL)
    header_row, cols = None, {}
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=6, max_col=8, values_only=True), 1):
        labels = [norm(c) for c in row]
        if any("cliente" in x for x in labels):
            for ci, lab in enumerate(labels, 1):
                if "fee" in lab:
                    cols["FEE"] = ci
                elif "variavel" in lab:
                    cols["VARIAVEL"] = ci
                elif "cliente" in lab and "cliente" not in cols:
                    cols["cliente"] = ci
            header_row = r
            break
    rows = []
    if header_row:
        cli_c = cols.get("cliente", 1)
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            nome = row[cli_c - 1] if cli_c - 1 < len(row) else None
            if not nome or not str(nome).strip():
                continue
            if norm(nome) in ("total", "totais", "soma"):
                continue
            for tipo, ci in (("FEE", cols.get("FEE")), ("VARIAVEL", cols.get("VARIAVEL"))):
                if not ci:
                    continue
                v = row[ci - 1] if ci - 1 < len(row) else None
                if isinstance(v, (int, float)) and v:
                    rows.append({"company": "REF", "year": year, "month": REF_MONTH,
                                 "cliente": str(nome).strip(), "tipo_receita": tipo,
                                 "value": round(float(v), 2), "source": src})
    wb.close()
    return pd.DataFrame(rows, columns=["company", "year", "month", "cliente", "tipo_receita", "value", "source"])
