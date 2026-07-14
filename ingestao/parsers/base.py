#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
parsers/base.py — lógica comum de leitura da aba DRE-Base (Technical Blueprint §6.2).

Cada empresa tem seu próprio módulo (ref.py, bd.py, ...) que apenas chama
parse_dre_base() com o código da empresa; aqui ficam as regras compartilhadas:
  - localização da aba e da linha de cabeçalho (datas datetime OU texto JAN/FEV...);
  - plano de contas CANÔNICO (mapeia variações de nomenclatura por empresa);
  - normalização (NaN->0.0, remoção de linhas vazias/%);
  - contrato do DataFrame de saída exigido pelo blueprint.

Contrato do DataFrame (exatamente estas colunas):
  company, year, month, account_code, account_description, group, value, source
"""
import os, unicodedata, datetime as dt
import pandas as pd
import openpyxl

MESES_PT = {1: "JANEIRO", 2: "FEVEREIRO", 3: "MARÇO", 4: "ABRIL", 5: "MAIO", 6: "JUNHO",
            7: "JULHO", 8: "AGOSTO", 9: "SETEMBRO", 10: "OUTUBRO", 11: "NOVEMBRO", 12: "DEZEMBRO"}
_TXT2MES = {}
for _i, _n in MESES_PT.items():
    _TXT2MES[_n] = _i
    _TXT2MES[_n[:3]] = _i
# inglês também (alguns arquivos podem usar)
for _i, _n in enumerate(["JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE", "JULY",
                         "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER"], 1):
    _TXT2MES[_n] = _i; _TXT2MES[_n[:3]] = _i

# Plano de contas CANÔNICO: (descricao, grupo, tipo, [(padrão_normalizado, modo)])
#   grupo ∈ REVENUE, DIRECT_COST, PERSONNEL, ADMIN, FACILITIES, FINANCIAL, TAXES, RESULT
#   modo: 'eq' (igualdade exata) | 'pre' (startswith). 1ª linha que casa (de cima p/ baixo) vence.
CANON = [
    ("RECEITA BRUTA",                                  "REVENUE",     "revenue", [("receita bruta", "eq")]),
    ("DEDUCOES IMPOSTOS",                              "TAXES",       "tax",     [("deducoes", "pre")]),
    ("RECEITA OPERACIONAL LIQUIDA",                    "REVENUE",     "revenue", [("receita operacional liquida", "eq")]),
    ("CUSTOS DOS SERVICOS",                            "DIRECT_COST", "cost",    [("custos dos servicos", "pre")]),
    ("RESULTADO OPERACIONAL DA AGENCIA",               "RESULT",      "result",  [("resultado op da agencia", "pre"),
                                                                                   ("resultado op agencia", "pre"),
                                                                                   ("resultado op da produtora", "pre"),
                                                                                   ("resultado op da empresa", "pre")]),
    ("GASTOS COM PESSOAL",                             "PERSONNEL",   "expense", [("gastos com pessoal", "eq")]),
    ("INFRAESTRUTURA",                                 "FACILITIES",  "expense", [("infra estrutura", "pre"), ("infraestrutura", "pre")]),
    ("OUTRAS DESPESAS",                                "ADMIN",       "expense", [("outras despesas", "eq")]),
    ("DESPESAS ADMINISTRATIVAS",                       "FINANCIAL",   "expense", [("despesas adm", "pre")]),
    ("TRIBUTOS FEDERAIS",                              "TAXES",       "tax",     [("tributos federais", "pre")]),
    ("RESULTADO OPERACIONAL ANTES DOS IMPOSTOS (EBIT)", "RESULT",     "result",  [("resultado operacional antes dos impostos", "eq")]),
    ("RESULTADO LIQUIDO",                              "RESULT",      "result",  [("resultado antes das participacoes", "pre"),
                                                                                   ("resultado liquido", "eq")]),
    ("GERACAO DE CAIXA",                               "RESULT",      "result",  [("geracao de caixa", "pre")]),
]


def norm(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    s = "".join(c.lower() if c.isalnum() else " " for c in s)
    return " ".join(s.split())


def _find_dre_sheet(wb):
    for nm in wb.sheetnames:
        if norm(nm).replace(" ", "") in ("drebase", "dre"):
            return wb[nm]
    for nm in wb.sheetnames:
        if "dre" in norm(nm):
            return wb[nm]
    return None


def _find_header(ws, year=2026):
    """Linha de cabeçalho; aceita datas datetime (2026-01-01) OU texto (JANEIRO/JAN).
    Retorna (row_idx, {col_idx: month_int})."""
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=12, max_col=24, values_only=True), 1):
        months = {}
        for ci, v in enumerate(row, 1):
            if isinstance(v, dt.datetime) and v.year == year:
                months[ci] = v.month
            elif isinstance(v, str):
                m = _TXT2MES.get(v.strip().upper())
                if m:
                    months[ci] = m
        if len(months) >= 6:
            return r, months
    return None, {}


def _match(text, pairs):
    n = norm(text)
    if not n:
        return False
    for p, mode in pairs:
        if mode == "eq" and n == p:
            return True
        if mode == "pre" and n.startswith(p):
            return True
    return False


def parse_dre_base(path, company, year=2026):
    """Lê a aba DRE-Base de `path` e devolve o DataFrame no contrato do blueprint
    (uma linha por conta canônica x mês). Carrega TODOS os 12 meses (fiel à planilha)."""
    src = os.path.basename(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = _find_dre_sheet(wb)
    if ws is None:
        wb.close()
        raise RuntimeError(f"{src}: aba DRE-Base não encontrada")
    hrow, months = _find_header(ws, year)
    if not months:
        wb.close()
        raise RuntimeError(f"{src}: cabeçalho de meses não encontrado")

    rows, taken = [], set()
    for row in ws.iter_rows(min_row=hrow + 1, max_row=ws.max_row, values_only=True):
        label = row[1] if len(row) > 1 else None          # coluna B
        if label is None or not str(label).strip():        # regra: ignorar linhas vazias
            continue
        for descricao, grupo, tipo, pats in CANON:
            if descricao in taken:
                continue
            if _match(label, pats):
                taken.add(descricao)
                for ci, mes in months.items():
                    v = row[ci - 1] if ci - 1 < len(row) else None
                    if isinstance(v, (int, float)):
                        val = round(float(v), 4)
                    elif tipo == "result":
                        # Linha de RESULTADO (RA/EBIT/Res.Líq.) com célula VAZIA ou #REF!
                        # (ex.: Zup 2026, cujo resultado é calculado por fórmula externa
                        # que está quebrada na origem) → NULL, não 0. Assim o cockpit
                        # mostra "—" (dado pendente) em vez de um falso "0" enganoso.
                        val = None
                    else:
                        val = 0.0                                     # receita/custo vazio -> 0
                    rows.append({"company": company, "year": year, "month": mes,
                                 "account_code": None, "account_description": descricao,
                                 "group": grupo, "value": val, "source": src})
                break
    wb.close()
    df = pd.DataFrame(rows, columns=["company", "year", "month", "account_code",
                                     "account_description", "group", "value", "source"])
    return df
