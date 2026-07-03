#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
tri_hist.py — ingestão do histórico TRIMESTRAL por empresa (API_CONTRACT.md §Iteração 2).

Fonte: aba de resumo trimestral de cada planilha (nome varia por empresa):
  REF='Comparativo tri 242526', BD/Viv='Resumo tri 242526',
  4PR='Resumo  tri 242526' (espaço duplo), Zup='Resumo tri 2526' (só 2 anos).
Layout: linha de cabeçalho com colunas '1º Trimestre24' / '2º Trimestres25' etc.
(+ colunas 'TOTAL 20xx' e '% S/ RECEITA BRUTA', ignoradas); linhas seguintes com o
nome da métrica na 1ª célula textual e os valores nas colunas dos trimestres.
Células de erro do Excel ('#REF!', '#ERROR!') são ignoradas (Zup 2026 tem várias).

Destino: fato_dre_tri_hist(empresa_id, ano, tri, metrica, valor) — CREATE TABLE IF NOT
EXISTS + upsert idempotente (ON CONFLICT DO UPDATE). Percentuais são gravados em PONTOS
percentuais (fração da planilha * 100 — convenção do contrato). Todos os anos da aba são
carregados, INCLUSIVE o corrente (duplicata de referência inofensiva: a API exclui o ano
pedido do hist[]).

Uso:  cd ingestao && python tri_hist.py
Env:  INCOMING (pasta das planilhas; default ../data/incoming)
"""
import os, sys, re, glob, pathlib, unicodedata

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from psycopg2.extras import execute_values

from db import get_conn, _empresa_id

ROOT = pathlib.Path(__file__).resolve().parent.parent
INCOMING = pathlib.Path(os.environ.get("INCOMING", ROOT / "data" / "incoming"))

# empresa -> prefixo do arquivo (mesmo padrão de main.py / folha_fees.py)
PREFIX = {"REF": "REF+", "BD": "BD", "4PR": "4PR", "VIV": "Viv", "ZUP": "Zup"}

# empresa -> nome da aba trimestral (comparado com espaços colapsados)
SHEET = {
    "REF": "Comparativo tri 242526",
    "BD":  "Resumo tri 242526",
    "VIV": "Resumo tri 242526",
    "4PR": "Resumo  tri 242526",   # espaço duplo no arquivo real
    "ZUP": "Resumo tri 2526",
}

# métricas em pontos percentuais (fração da planilha * 100)
PCT_METRICAS = {"EBIT_NEG_PCT", "EBIT_AG_PCT"}

DDL = """
CREATE TABLE IF NOT EXISTS fato_dre_tri_hist (
    id         SERIAL PRIMARY KEY,
    empresa_id INT NOT NULL REFERENCES dim_empresa(id),
    ano        INT NOT NULL,
    tri        INT NOT NULL,
    metrica    VARCHAR(60) NOT NULL,
    valor      NUMERIC(16,2),
    UNIQUE (empresa_id, ano, tri, metrica)
);
"""


def _find_file(prefix):
    hits = sorted(glob.glob(str(INCOMING / f"{prefix}*DRE*.xlsx")))
    return hits[0] if hits else None


def _norm(s):
    """Maiúsculas sem acento e com espaços colapsados (planilhas têm encoding variado)."""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", s).strip().upper()


def _metrica(label):
    """Mapeia o rótulo da linha da planilha -> métrica canônica (ou None = ignora).
    Aliases por empresa: BD usa 'PRODUTORA' onde as agências usam 'AGENCIA'."""
    t = _norm(label)
    if t == "RECEITA BRUTA":
        return "RECEITA_BRUTA"
    if t.startswith("RECEITA OPERACIONAL"):
        return "RECEITA_LIQUIDA"
    if t.startswith("RESULTADO OP. DA"):                # AGENCIA (REF/Viv/4PR/Zup) / PRODUTORA (BD)
        return "RESULTADO_AGENCIA"
    if t.startswith("RESULTADO OPERACIONAL ANTES"):     # '... DOS IMPOSTOS' / '... IR E CSLL'
        return "EBIT"                                   # (a variante 'RESULTADO OP. ANTES ...' é duplicada — ignorada)
    if t.startswith("% EBIT NEG"):
        return "EBIT_NEG_PCT"
    if t.startswith("% EBIT AG") or t.startswith("% EBIT PRODUT"):
        return "EBIT_AG_PCT"
    if t == "RESULTADO LIQUIDO":
        return "RESULTADO_LIQUIDO"
    return None


# '1º Trimestre24' / '2º Trimestres25' -> (tri, ano). Colunas TOTAL/% não casam.
_RE_TRI = re.compile(r"^([1-4])\s*\S*\s*TRIMESTRES?\s*(\d{2})$")


def _header_cols(ws):
    """Encontra a linha de cabeçalho e devolve {idx_coluna: (ano, tri)}."""
    for row in ws.iter_rows(min_row=1, max_row=6, max_col=30, values_only=True):
        cols = {}
        for i, v in enumerate(row):
            if not isinstance(v, str):
                continue
            m = _RE_TRI.match(_norm(v))
            if m:
                cols[i] = (2000 + int(m.group(2)), int(m.group(1)))
        if cols:
            return cols
    return {}


def parse_tri_hist(path, empresa):
    """Lê a aba trimestral -> [(ano, tri, metrica, valor)] (pct já em pontos)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        alvo = _norm(SHEET[empresa])
        nm = next((s for s in wb.sheetnames if _norm(s) == alvo), None)
        if not nm:
            raise ValueError(f"aba trimestral não encontrada ({SHEET[empresa]!r})")
        ws = wb[nm]
        cols = _header_cols(ws)
        if not cols:
            raise ValueError(f"cabeçalho de trimestres não encontrado na aba {nm!r}")
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=60, max_col=30, values_only=True):
            label = next((v for v in row if isinstance(v, str) and v.strip()), None)
            met = _metrica(label) if label else None
            if not met:
                continue
            for i, (ano, tri) in cols.items():
                v = row[i] if i < len(row) else None
                if not isinstance(v, (int, float)) or isinstance(v, bool):
                    continue                     # '#REF!'/'#ERROR!'/vazio -> ignora
                val = float(v) * 100.0 if met in PCT_METRICAS else float(v)
                rows.append((ano, tri, met, round(val, 2)))
        return rows
    finally:
        wb.close()


def upsert_tri_hist(cur, empresa, rows, ec):
    """Upsert idempotente em fato_dre_tri_hist. Retorna nº de linhas."""
    if not rows:
        return 0
    emp_id = _empresa_id(cur, empresa, ec)
    execute_values(cur, """
        INSERT INTO fato_dre_tri_hist (empresa_id, ano, tri, metrica, valor) VALUES %s
        ON CONFLICT (empresa_id, ano, tri, metrica) DO UPDATE SET valor = EXCLUDED.valor
    """, [(emp_id, a, t, m, v) for a, t, m, v in rows])
    return len(rows)


def run():
    """Executa a ingestão IN-PROCESS e devolve {ok, output} (padrão de folha_fees.run())."""
    out = [f"== Histórico trimestral (fato_dre_tri_hist) == (planilhas em {INCOMING})"]
    ok = True
    con = get_conn(); con.autocommit = False
    try:
        cur = con.cursor()
        cur.execute(DDL)                      # tabela do API_CONTRACT.md (idempotente)
        con.commit()
        ec = {}
        for empresa, prefix in PREFIX.items():
            caminho = _find_file(prefix)
            if not caminho:
                out.append(f"  [{empresa}] arquivo não encontrado (prefixo {prefix})")
                ok = False
                continue
            try:
                rows = parse_tri_hist(caminho, empresa)
                n = upsert_tri_hist(cur, empresa, rows, ec)
                con.commit()
                anos = sorted({a for a, _, _, _ in rows})
                out.append(f"  [{empresa}] OK: {n} linhas (anos {anos})")
                if n == 0:
                    out.append(f"  [{empresa}] ALERTA: nenhuma linha trimestral")
                    ok = False
            except Exception as e:
                con.rollback()
                out.append(f"  [{empresa}] ERRO: {e}")
                ok = False
    finally:
        con.close()
    return {"ok": ok, "output": "\n".join(out)}


def main():
    r = run()
    print(r["output"])
    sys.exit(0 if r["ok"] else 1)


if __name__ == "__main__":
    main()
