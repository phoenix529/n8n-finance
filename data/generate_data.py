#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Cockpit Financeiro Estratégico — Gerador de dados de exemplo (determinístico).

Parte do projeto "Cockpit Financeiro Estratégico" (n8n + PostgreSQL 16 + pgvector + RAG).
Este script é a ÚNICA fonte dos dados sintéticos usados por toda a stack:

  * data/out/fact_financials.csv ........ tabela fato longa (carga no Postgres)
  * data/raw/historico_<company_id>.xlsx . planilhas históricas por empresa (ingestão)
  * data/raw/upload_AUR-VAR_2026-06.xlsx . upload "limpo" do mês corrente (demo)
  * data/raw/upload_AUR-IND_2026-06_INVALIDO.xlsx . upload malformado (validação/quarentena)
  * data/out/dashboard_data.json ......... contrato do dashboard (seção 8 do SPEC)
  * dashboard/dashboard_data.json ........ cópia idêntica lida pelo front-end

Conformidade total com SPEC.md (fonte canônica): nomes de empresas, account_code,
fórmulas de KPI, períodos e o formato do JSON do dashboard.

Requisitos: Python 3.13, APENAS biblioteca padrão. openpyxl é OPCIONAL — se ausente,
geramos apenas os .csv de fallback e seguimos sem quebrar.

Determinístico: random.seed(42). Rodar `python data/generate_data.py` reproduz tudo.

NÃO executa nada destrutivo; apenas escreve arquivos sob data/ e dashboard/.
"""

from __future__ import annotations

import csv
import json
import math
import os
import random
from datetime import date

# --------------------------------------------------------------------------------------
# 0. Import opcional e guardado de openpyxl
# --------------------------------------------------------------------------------------
try:
    import openpyxl  # type: ignore
    from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore
    _HAS_OPENPYXL = True
except Exception:  # pragma: no cover - ambiente sem openpyxl
    openpyxl = None  # type: ignore
    _HAS_OPENPYXL = False

# --------------------------------------------------------------------------------------
# 1. Determinismo
# --------------------------------------------------------------------------------------
SEED = 42
random.seed(SEED)

# --------------------------------------------------------------------------------------
# 2. Caminhos (absolutos no Windows, conforme SPEC seção 9)
# --------------------------------------------------------------------------------------
PROJECT_ROOT = r"C:\Users\Administrator\Documents\n8n"
DATA_DIR = os.path.join(PROJECT_ROOT, "data")
OUT_DIR = os.path.join(DATA_DIR, "out")
RAW_DIR = os.path.join(DATA_DIR, "raw")
DASHBOARD_DIR = os.path.join(PROJECT_ROOT, "dashboard")

FACT_CSV = os.path.join(OUT_DIR, "fact_financials.csv")
DASHBOARD_JSON_OUT = os.path.join(OUT_DIR, "dashboard_data.json")
DASHBOARD_JSON_COPY = os.path.join(DASHBOARD_DIR, "dashboard_data.json")

# --------------------------------------------------------------------------------------
# 3. Dimensões — empresas (SPEC seção 2) e plano de contas (SPEC seção 3)
# --------------------------------------------------------------------------------------
COMPANIES = [
    # company_id, name, sector, color, receita_anual_aprox (R$), is_consolidating, sort
    ("AUR-VAR", "Aurora Varejo S.A.",        "Varejo",     "#4F6BED", 210_000_000, True, 1),
    ("AUR-IND", "Aurora Indústria Ltda.",    "Indústria",  "#0EA5E9", 150_000_000, True, 2),
    ("AUR-SVC", "Aurora Serviços Ltda.",     "Serviços",   "#10B981",  70_000_000, True, 3),
    ("AUR-LOG", "Aurora Logística Ltda.",    "Logística",  "#F59E0B",  50_000_000, True, 4),
    ("AUR-HLD", "Aurora Participações S.A.", "Holding",    "#6B7280",     500_000, True, 5),
    ("ELIM",    "Eliminações Intercompany",  "Eliminação", "#94A3B8",           0, True, 6),
]
COMPANY_BY_ID = {c[0]: c for c in COMPANIES}

# Plano de contas P&L (account_kind = 'PNL'). Sinal: receita +, custos/despesas negativos.
PNL_ACCOUNTS = [
    # account_code, account_name, group_code, sign
    ("R_BRUTA",      "Receita Bruta de Vendas",            "RECEITA", +1),
    ("DEDUCOES",     "Impostos e Deduções s/ Vendas",      "RECEITA", -1),
    ("CMV",          "Custo dos Produtos/Serviços (CMV)",  "CUSTO",   -1),
    ("DESP_PESSOAL", "Despesas com Pessoal",               "OPEX",    -1),
    ("DESP_VENDAS",  "Despesas Comerciais e Marketing",    "OPEX",    -1),
    ("DESP_ADM",     "Despesas Administrativas",           "OPEX",    -1),
    ("DESP_OUTRAS",  "Outras Despesas Operacionais",       "OPEX",    -1),
    ("DEPRECIACAO",  "Depreciação e Amortização",          "DA",      -1),
    ("RESULT_FIN",   "Resultado Financeiro Líquido",       "FINANC",  +1),
    ("IRPJ_CSLL",    "IR e CSLL",                          "IMPOSTO", -1),
]

# Plano de contas de posição (account_kind = 'POSICAO'), todos positivos.
POSICAO_ACCOUNTS = [
    ("CAIXA",      "Caixa e Equivalentes"),
    ("AR",         "Contas a Receber"),
    ("AP",         "Contas a Pagar"),
    ("ESTOQUE",    "Estoques"),
    ("DIVIDA",     "Dívida Bruta (Empréstimos)"),
    ("PATRIMONIO", "Patrimônio Líquido"),
]

ACCOUNT_NAME = {c: n for c, n, *_ in PNL_ACCOUNTS}
ACCOUNT_NAME.update({c: n for c, n in POSICAO_ACCOUNTS})
VALID_ACCOUNT_CODES = set(ACCOUNT_NAME.keys())

# Códigos P&L em ordem canônica para iteração estável
PNL_CODES = [c for c, *_ in PNL_ACCOUNTS]
POSICAO_CODES = [c for c, _ in POSICAO_ACCOUNTS]

# --------------------------------------------------------------------------------------
# 4. Períodos (SPEC seção 5): 2025-01 .. 2026-06 (mensal, primeiro dia do mês)
# --------------------------------------------------------------------------------------
def month_range(start: date, end: date):
    """Gera o primeiro dia de cada mês de start (inclusive) até end (inclusive)."""
    y, m = start.year, start.month
    out = []
    while (y, m) <= (end.year, end.month):
        out.append(date(y, m, 1))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return out

PERIOD_START = date(2025, 1, 1)
PERIOD_END = date(2026, 6, 1)
LAST_CLOSED = date(2026, 5, 1)          # SPEC: last_closed_period
PERIODS = month_range(PERIOD_START, PERIOD_END)      # 18 meses (17 fechados + jun/26 parcial)
GENERATED_AT = "2026-06-23"

def ym(d: date) -> str:
    return f"{d.year:04d}-{d.month:02d}"

PERIOD_YM = [ym(d) for d in PERIODS]

# --------------------------------------------------------------------------------------
# 5. Perfis setoriais — margens, sazonalidade, crescimento
# --------------------------------------------------------------------------------------
# Cada perfil descreve a estrutura econômica típica do setor (frações da receita líquida).
# Sinais aplicados depois; aqui guardamos magnitudes positivas como fração da RECEITA LÍQUIDA.
SECTOR_PROFILE = {
    "Varejo": {
        "ded_frac": 0.255,        # deduções sobre receita bruta
        "cmv_frac": 0.620,        # CMV alto (revenda)
        "pessoal_frac": 0.110,
        "vendas_frac": 0.075,
        "adm_frac": 0.040,
        "outras_frac": 0.012,
        "deprec_frac": 0.018,
        "fin_frac": -0.020,       # resultado financeiro negativo (dívida)
        "ir_aliq": 0.34,          # alíquota efetiva sobre LAIR positivo
        "growth_yoy": 0.085,      # +8,5% a.a.
        "base_receita_liq_mensal": 14_300_000,   # ~ R$ 171,6M líquida/ano antes da sazonalidade
    },
    "Indústria": {
        "ded_frac": 0.230,
        "cmv_frac": 0.560,
        "pessoal_frac": 0.130,
        "vendas_frac": 0.045,
        "adm_frac": 0.050,
        "outras_frac": 0.015,
        "deprec_frac": 0.050,     # capital intensivo
        "fin_frac": -0.028,
        "ir_aliq": 0.34,
        "growth_yoy": 0.060,
        "base_receita_liq_mensal": 9_600_000,
    },
    "Serviços": {
        "ded_frac": 0.160,        # ISS menor
        "cmv_frac": 0.280,        # custo de serviço (mão de obra direta)
        "pessoal_frac": 0.300,    # intensivo em pessoal
        "vendas_frac": 0.060,
        "adm_frac": 0.070,
        "outras_frac": 0.018,
        "deprec_frac": 0.020,
        "fin_frac": -0.010,
        "ir_aliq": 0.34,
        "growth_yoy": 0.140,      # serviços crescendo forte
        "base_receita_liq_mensal": 4_900_000,
    },
    "Logística": {
        "ded_frac": 0.180,
        "cmv_frac": 0.520,        # combustível, frete terceirizado
        "pessoal_frac": 0.160,
        "vendas_frac": 0.030,
        "adm_frac": 0.055,
        "outras_frac": 0.020,
        "deprec_frac": 0.060,     # frota
        "fin_frac": -0.035,       # leasing/financiamento de frota
        "ir_aliq": 0.34,
        "growth_yoy": 0.075,
        "base_receita_liq_mensal": 3_500_000,
    },
    "Holding": {
        # AUR-HLD: corporativo. Receita ínfima (rateio de serviços), custos de holding.
        "ded_frac": 0.050,
        "cmv_frac": 0.000,
        "pessoal_frac": 1.10,     # custos corporativos > receita -> EBITDA negativo
        "vendas_frac": 0.05,
        "adm_frac": 0.90,
        "outras_frac": 0.10,
        "deprec_frac": 0.08,
        "fin_frac": 0.40,         # holding capta dividendos/juros -> financeiro positivo
        "ir_aliq": 0.10,
        "growth_yoy": 0.030,
        "base_receita_liq_mensal": 35_000,        # ~R$ 420k/ano
    },
    "Eliminação": None,  # ELIM ~ 0 (architecture-ready)
}

# Fatores de sazonalidade por mês do ano (índice 1..12). Pico de varejo em Nov/Dez.
# Aplicados sobre a receita; demais setores usam um perfil mais suave.
SEASONALITY_VAREJO = {
    1: 0.86, 2: 0.84, 3: 0.95, 4: 0.97, 5: 1.00, 6: 0.98,
    7: 0.99, 8: 1.01, 9: 1.03, 10: 1.08, 11: 1.22, 12: 1.37,
}
SEASONALITY_GERAL = {
    1: 0.94, 2: 0.92, 3: 1.00, 4: 1.01, 5: 1.02, 6: 1.00,
    7: 0.99, 8: 1.02, 9: 1.03, 10: 1.04, 11: 1.05, 12: 1.06,
}
SEASONALITY_SERVICOS = {
    1: 0.95, 2: 0.93, 3: 1.00, 4: 1.01, 5: 1.02, 6: 1.01,
    7: 1.00, 8: 1.02, 9: 1.03, 10: 1.05, 11: 1.04, 12: 0.98,  # cai um pouco em dez (férias)
}

def seasonality_factor(sector: str, month: int) -> float:
    if sector == "Varejo":
        return SEASONALITY_VAREJO[month]
    if sector == "Serviços":
        return SEASONALITY_SERVICOS[month]
    return SEASONALITY_GERAL[month]

# --------------------------------------------------------------------------------------
# 6. Geração do P&L por empresa-mês
# --------------------------------------------------------------------------------------
def money(x: float) -> float:
    """Arredonda para centavos (2 casas)."""
    return round(float(x), 2)

def jitter(pct: float) -> float:
    """Ruído multiplicativo determinístico em +/- pct."""
    return 1.0 + random.uniform(-pct, pct)

def months_from_start(d: date) -> int:
    return (d.year - PERIOD_START.year) * 12 + (d.month - PERIOD_START.month)

def growth_factor(profile: dict, d: date) -> float:
    """Crescimento composto mensal derivado do YoY do setor."""
    monthly_g = (1.0 + profile["growth_yoy"]) ** (1.0 / 12.0)
    return monthly_g ** months_from_start(d)

def build_pnl_company_month(company_id: str, d: date) -> dict:
    """
    Retorna {account_code: valor_realizado} para as contas P&L da empresa no mês.
    Sinais já aplicados conforme convenção (receita +, custos/despesas -).
    ELIM -> tudo ~0.
    """
    sector = COMPANY_BY_ID[company_id][2]
    if sector == "Eliminação":
        # ELIM ~ 0: pequenos valores simétricos (arquitetura pronta p/ eliminações reais).
        elim = {code: 0.0 for code in PNL_CODES}
        # valores residuais minúsculos para mostrar que a linha existe (ainda ~0 no consolidado)
        tiny = round(random.uniform(-1500, 1500), 2)
        elim["R_BRUTA"] = abs(tiny)
        elim["DEDUCOES"] = -abs(tiny) * 0.1
        elim["CMV"] = -abs(tiny) * 0.4
        elim["DESP_ADM"] = -abs(tiny) * 0.2
        return {k: money(v) for k, v in elim.items()}

    p = SECTOR_PROFILE[sector]

    # Receita líquida base com crescimento + sazonalidade + ruído
    base_liq = p["base_receita_liq_mensal"]
    rl = base_liq * growth_factor(p, d) * seasonality_factor(sector, d.month) * jitter(0.03)

    # Receita bruta a partir da líquida (RL = RB + DEDUCOES, DEDUCOES negativo)
    ded_frac = p["ded_frac"]
    r_bruta = rl / (1.0 - ded_frac)
    deducoes = -(r_bruta - rl)

    # Custos/despesas como fração da receita líquida (com leve ruído por linha)
    cmv = -rl * p["cmv_frac"] * jitter(0.02)
    desp_pessoal = -rl * p["pessoal_frac"] * jitter(0.015)
    desp_vendas = -rl * p["vendas_frac"] * jitter(0.04)
    desp_adm = -rl * p["adm_frac"] * jitter(0.02)
    desp_outras = -rl * p["outras_frac"] * jitter(0.06)
    deprec = -rl * p["deprec_frac"] * jitter(0.005)  # depreciação estável

    # Resultado financeiro (proporcional à receita; sinal vem do perfil)
    result_fin = rl * p["fin_frac"] * jitter(0.05)

    # EBIT preliminar para estimar IR
    ebitda = rl + cmv + desp_pessoal + desp_vendas + desp_adm + desp_outras
    ebit = ebitda + deprec
    lair = ebit + result_fin  # lucro antes do IR
    irpj = -max(lair, 0.0) * p["ir_aliq"]

    vals = {
        "R_BRUTA": r_bruta,
        "DEDUCOES": deducoes,
        "CMV": cmv,
        "DESP_PESSOAL": desp_pessoal,
        "DESP_VENDAS": desp_vendas,
        "DESP_ADM": desp_adm,
        "DESP_OUTRAS": desp_outras,
        "DEPRECIACAO": deprec,
        "RESULT_FIN": result_fin,
        "IRPJ_CSLL": irpj,
    }
    return {k: money(v) for k, v in vals.items()}

# --------------------------------------------------------------------------------------
# 7. Geração das contas de POSIÇÃO (balanço/caixa) por empresa-mês — evolução coerente
# --------------------------------------------------------------------------------------
def build_position_series(company_id: str, pnl_by_period: dict) -> dict:
    """
    Constrói séries de posição (estoque, AR, AP, dívida, caixa, patrimônio) coerentes
    com o P&L. Retorna {period_ym: {account_code: valor}}.

    Modelo simplificado mas consistente:
      AR      ~ ligado à receita bruta (DSO setorial)
      ESTOQUE ~ ligado ao CMV (giro setorial); ~0 para serviços/holding
      AP      ~ ligado a (CMV + despesas) (prazo de pagamento)
      DIVIDA  ~ nível por empresa, amortização/captação suave
      CAIXA   ~ caixa(m-1) + fluxo operacional aproximado + variação de capital de giro
                - amortização de dívida + financeiro
      PATRIMONIO ~ acumula lucro líquido (proxy), partindo de uma base
    """
    sector = COMPANY_BY_ID[company_id][2]
    out = {}

    if sector == "Eliminação":
        for ymk in PERIOD_YM:
            out[ymk] = {code: 0.0 for code in POSICAO_CODES}
        return out

    p = SECTOR_PROFILE[sector]

    # Parâmetros setoriais de prazos (dias) e estoque
    dso_dias = {"Varejo": 22, "Indústria": 48, "Serviços": 38, "Logística": 41, "Holding": 30}[sector]
    dio_dias = {"Varejo": 55, "Indústria": 70, "Serviços": 4, "Logística": 6, "Holding": 0}[sector]  # estoque
    dpo_dias = {"Varejo": 40, "Indústria": 52, "Serviços": 30, "Logística": 33, "Holding": 25}[sector]

    # Dívida bruta inicial e patrimônio inicial por empresa (R$)
    divida0 = {
        "AUR-VAR": 78_000_000, "AUR-IND": 92_000_000, "AUR-SVC": 12_000_000,
        "AUR-LOG": 31_000_000, "AUR-HLD": 6_000_000,
    }[company_id]
    patrimonio0 = {
        "AUR-VAR": 95_000_000, "AUR-IND": 130_000_000, "AUR-SVC": 28_000_000,
        "AUR-LOG": 22_000_000, "AUR-HLD": 40_000_000,
    }[company_id]
    caixa0 = {
        "AUR-VAR": 24_000_000, "AUR-IND": 18_000_000, "AUR-SVC": 9_000_000,
        "AUR-LOG": 5_500_000, "AUR-HLD": 12_000_000,
    }[company_id]

    prev_caixa = caixa0
    prev_patrimonio = patrimonio0
    prev_divida = divida0
    prev_wc = None  # capital de giro anterior (AR + ESTOQUE - AP)

    for idx, ymk in enumerate(PERIOD_YM):
        pnl = pnl_by_period[ymk]
        r_bruta = pnl["R_BRUTA"]
        cmv_abs = abs(pnl["CMV"])
        desp_abs = (abs(pnl["DESP_PESSOAL"]) + abs(pnl["DESP_VENDAS"]) +
                    abs(pnl["DESP_ADM"]) + abs(pnl["DESP_OUTRAS"]))

        # Contas a Receber a partir do DSO: AR = R_BRUTA / 30 * DSO
        ar = r_bruta / 30.0 * dso_dias * jitter(0.03)
        # Estoque a partir do CMV mensal: ESTOQUE = CMV / 30 * DIO
        estoque = cmv_abs / 30.0 * dio_dias * jitter(0.04) if dio_dias > 0 else 0.0
        # Contas a Pagar: sobre custos+despesas
        ap = (cmv_abs + desp_abs) / 30.0 * dpo_dias * jitter(0.03)

        # Dívida: amortização leve com sazonal captação no início do ano
        amort = prev_divida * 0.006  # ~0,6% a.m.
        capta = prev_divida * 0.010 if ymk.endswith("-01") else 0.0  # captação em janeiro
        divida = max(prev_divida - amort + capta, 0.0) * jitter(0.002)

        # Lucro líquido do mês (proxy de geração de caixa para patrimônio e caixa)
        receita_liq = pnl["R_BRUTA"] + pnl["DEDUCOES"]
        ebitda = (receita_liq + pnl["CMV"] + pnl["DESP_PESSOAL"] + pnl["DESP_VENDAS"] +
                  pnl["DESP_ADM"] + pnl["DESP_OUTRAS"])
        ebit = ebitda + pnl["DEPRECIACAO"]
        lucro_liq = ebit + pnl["RESULT_FIN"] + pnl["IRPJ_CSLL"]

        # Variação de capital de giro (consome/gera caixa)
        wc = ar + estoque - ap
        delta_wc = 0.0 if prev_wc is None else (wc - prev_wc)

        # Fluxo de caixa do mês (proxy): EBITDA - var.WC - amort + capta + financeiro - IR
        # (depreciação não é caixa). Dividendos tratados abaixo.
        fc = (ebitda
              - delta_wc
              - amort + capta
              + pnl["RESULT_FIN"]
              + pnl["IRPJ_CSLL"])
        # CAPEX aproximado: parte da depreciação reinvestida
        capex = -abs(pnl["DEPRECIACAO"]) * 0.9
        fc += capex
        # Dividendos/juros sobre capital: holding distribui menos caixa (retém)
        if sector != "Holding":
            fc -= max(lucro_liq, 0.0) * 0.15  # paga ~15% do lucro como dividendo

        caixa = max(prev_caixa + fc, 50_000.0)  # piso de caixa para evitar negativos absurdos

        # Patrimônio acumula lucro retido
        patrimonio = prev_patrimonio + lucro_liq * 0.85

        out[ymk] = {
            "CAIXA": money(caixa),
            "AR": money(ar),
            "AP": money(ap),
            "ESTOQUE": money(estoque),
            "DIVIDA": money(divida),
            "PATRIMONIO": money(patrimonio),
        }

        prev_caixa = caixa
        prev_patrimonio = patrimonio
        prev_divida = divida
        prev_wc = wc

    return out

# --------------------------------------------------------------------------------------
# 8. Orçamento (valor_orcado) — dentro de +/-8% do realizado, com variâncias deliberadas
# --------------------------------------------------------------------------------------
# Algumas (empresa, conta) recebem variância "estourada" para a demo de Budget-vs-Actual.
#   Convenção: orcado = realizado / fator  =>  var% (realizado vs orçado) = fator - 1.
#   Receita: fator>1 = receita acima do orçado (favorável); fator<1 = receita abaixo (miss).
#   Despesa (realizado negativo): fator>1 = estouro de gasto (desfavorável); fator<1 = economia.
#   Mix deliberado de itens favoráveis E desfavoráveis que se compensam na consolidação,
#   evitando um lucro líquido irrealisticamente acima do orçado.
DELIBERATE_VARIANCES = {
    # (company_id, account_code): fator
    ("AUR-VAR", "R_BRUTA"):     1.03,   # Varejo: receita 3% acima do orçado (favorável)
    ("AUR-SVC", "R_BRUTA"):     1.05,   # Serviços: receita 5% acima do orçado (favorável)
    ("AUR-IND", "R_BRUTA"):     0.98,   # Indústria: receita 2% abaixo do orçado (miss)
    ("AUR-VAR", "DESP_VENDAS"): 1.16,   # Varejo: marketing estourou 16% acima do orçado (alerta)
    ("AUR-IND", "CMV"):         1.05,   # Indústria: CMV 5% acima do orçado (insumo subiu)
    ("AUR-LOG", "DESP_OUTRAS"): 1.10,   # Logística: manutenção de frota estourou 10%
}

def _stable_unit(s: str) -> float:
    """Hash determinístico estável (independente de PYTHONHASHSEED) -> float em [0,1)."""
    h = 0
    for ch in s:
        h = (h * 131 + ord(ch)) % 1_000_003
    return h / 1_000_003


def budget_for(company_id: str, account_code: str, realizado: float, ymk: str) -> float:
    """
    Orçamento COERENTE por (empresa, período): um viés único de receita aplicado a TODAS as
    linhas (preserva margens) + jitter mínimo por linha. Isso evita que o lucro líquido — um
    resíduo pequeno de números grandes — exploda em variância quando cada linha é perturbada de
    forma independente. Algumas (empresa, conta) recebem variância deliberada (estouro/beat) para
    popular a tela de Orçado-vs-Realizado e os alertas. Para realizado ~0 (ELIM), orçado ~0.
    """
    if abs(realizado) < 1.0:
        return money(realizado)

    key = (company_id, account_code)
    if key in DELIBERATE_VARIANCES:
        factor = DELIBERATE_VARIANCES[key]
        # orcado = realizado / factor  -> variacao_pct = realizado/orcado - 1 = factor - 1
        return money(realizado / factor)

    # viés comum da empresa no período (~ +/-3%) -> margem preservada, var. de fundo modesta
    bias = (_stable_unit(company_id + "|" + ymk) - 0.5) * 2 * 0.03
    # jitter por linha (~ +/-1%) -> pequena dispersão linha-a-linha, sem distorcer o resíduo
    jitter = (_stable_unit(company_id + "|" + ymk + "|" + account_code) - 0.5) * 2 * 0.01
    orcado = realizado / (1.0 + bias + jitter)
    return money(orcado)

# --------------------------------------------------------------------------------------
# 9. Montagem da tabela fato completa (realizado + orçado, P&L + posição)
# --------------------------------------------------------------------------------------
LOAD_ID_BASE = "seed-2026-06-23"

def source_file_for(company_id: str) -> str:
    return f"historico_{company_id}.xlsx"

def build_all_facts():
    """
    Retorna:
      facts: lista de dicts {company_id, period_date, account_code, valor_realizado,
                             valor_orcado, source_file, load_id}
      pnl: dict[company_id][period_ym][account_code] = realizado
      pos: dict[company_id][period_ym][account_code] = realizado
    """
    facts = []
    pnl_all = {}
    pos_all = {}

    for (company_id, *_rest) in COMPANIES:
        # 9.1 P&L por período
        pnl_by_period = {}
        for d in PERIODS:
            pnl_by_period[ym(d)] = build_pnl_company_month(company_id, d)
        pnl_all[company_id] = pnl_by_period

        # 9.2 Posição por período (depende do P&L)
        pos_by_period = build_position_series(company_id, pnl_by_period)
        pos_all[company_id] = pos_by_period

        src = source_file_for(company_id)
        load_id = f"{LOAD_ID_BASE}-{company_id}"

        # 9.3 Linhas P&L
        for d in PERIODS:
            ymk = ym(d)
            for code in PNL_CODES:
                realizado = pnl_by_period[ymk][code]
                orcado = budget_for(company_id, code, realizado, ymk)
                facts.append({
                    "company_id": company_id,
                    "period_date": d.isoformat(),
                    "account_code": code,
                    "valor_realizado": money(realizado),
                    "valor_orcado": money(orcado),
                    "source_file": src,
                    "load_id": load_id,
                })
            # 9.4 Linhas de posição
            for code in POSICAO_CODES:
                realizado = pos_by_period[ymk][code]
                # orçado de posição: meta de balanço coerente por (empresa, período)
                orcado = budget_for(company_id, code, realizado, ymk)
                facts.append({
                    "company_id": company_id,
                    "period_date": d.isoformat(),
                    "account_code": code,
                    "valor_realizado": money(realizado),
                    "valor_orcado": money(orcado),
                    "source_file": src,
                    "load_id": load_id,
                })

    return facts, pnl_all, pos_all

# --------------------------------------------------------------------------------------
# 10. Derivação de KPIs (SPEC seção 4) — por empresa-mês e consolidado
# --------------------------------------------------------------------------------------
def derive_pnl(account_map: dict) -> dict:
    """A partir de {account_code: valor} (realizado), deriva KPIs de P&L (SPEC seção 4)."""
    g = lambda c: account_map.get(c, 0.0)
    receita_liquida = g("R_BRUTA") + g("DEDUCOES")
    lucro_bruto = receita_liquida + g("CMV")
    ebitda = lucro_bruto + g("DESP_PESSOAL") + g("DESP_VENDAS") + g("DESP_ADM") + g("DESP_OUTRAS")
    ebit = ebitda + g("DEPRECIACAO")
    lucro_liquido = ebit + g("RESULT_FIN") + g("IRPJ_CSLL")
    mb = (lucro_bruto / receita_liquida * 100.0) if receita_liquida else 0.0
    me = (ebitda / receita_liquida * 100.0) if receita_liquida else 0.0
    ml = (lucro_liquido / receita_liquida * 100.0) if receita_liquida else 0.0
    return {
        "receita_bruta": g("R_BRUTA"),
        "receita_liquida": receita_liquida,
        "lucro_bruto": lucro_bruto,
        "ebitda": ebitda,
        "ebit": ebit,
        "lucro_liquido": lucro_liquido,
        "margem_bruta_pct": mb,
        "margem_ebitda_pct": me,
        "margem_liquida_pct": ml,
    }

def derive_position(account_map: dict, ebitda_ltm: float | None = None) -> dict:
    g = lambda c: account_map.get(c, 0.0)
    divida_liquida = g("DIVIDA") - g("CAIXA")
    capital_giro = g("AR") + g("ESTOQUE") - g("AP")
    dso = (g("AR") / g("R_BRUTA") * 30.0) if g("R_BRUTA") else 0.0
    out = {
        "caixa": g("CAIXA"),
        "ar": g("AR"),
        "ap": g("AP"),
        "estoque": g("ESTOQUE"),
        "divida": g("DIVIDA"),
        "patrimonio": g("PATRIMONIO"),
        "divida_liquida": divida_liquida,
        "capital_giro": capital_giro,
        "dso_dias": dso,
    }
    if ebitda_ltm:
        out["divida_ebitda"] = divida_liquida / ebitda_ltm if ebitda_ltm else 0.0
    return out

def consolidate_period(pnl_all: dict, pos_all: dict, ymk: str) -> dict:
    """Soma contas-folha de todas as empresas (consolidado) e deriva KPIs do mês."""
    summed_pnl = {code: 0.0 for code in PNL_CODES}
    summed_pos = {code: 0.0 for code in POSICAO_CODES}
    for company_id, *_ in COMPANIES:
        for code in PNL_CODES:
            summed_pnl[code] += pnl_all[company_id][ymk][code]
        for code in POSICAO_CODES:
            summed_pos[code] += pos_all[company_id][ymk][code]
    derived = derive_pnl(summed_pnl)
    # adiciona R_BRUTA para cálculo de DSO consolidado
    merged = dict(summed_pos)
    merged["R_BRUTA"] = summed_pnl["R_BRUTA"]
    pos_derived = derive_position(merged)
    return {"pnl_leaf": summed_pnl, "pos_leaf": summed_pos,
            "pnl": derived, "pos": pos_derived}

# --------------------------------------------------------------------------------------
# 11. Escrita do fact_financials.csv (SPEC seção 2 do prompt / seção 9 do SPEC)
# --------------------------------------------------------------------------------------
def write_fact_csv(facts: list) -> int:
    os.makedirs(OUT_DIR, exist_ok=True)
    cols = ["company_id", "period_date", "account_code",
            "valor_realizado", "valor_orcado", "source_file", "load_id"]
    with open(FACT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for row in facts:
            w.writerow(row)
    return len(facts)

# --------------------------------------------------------------------------------------
# 12. Planilhas de ingestão por empresa (shape de upload real) + demos
# --------------------------------------------------------------------------------------
# Colunas do upload (PT-BR, como uma planilha de área financeira):
UPLOAD_COLS = ["empresa", "periodo", "conta_codigo", "conta_nome",
               "valor_realizado", "valor_orcado"]

def upload_rows_for_company(company_id: str, facts: list) -> list:
    """Linhas no formato de upload para uma empresa (todos os períodos e contas)."""
    name = COMPANY_BY_ID[company_id][1]
    rows = []
    for fr in facts:
        if fr["company_id"] != company_id:
            continue
        period_ym = fr["period_date"][:7]  # YYYY-MM
        rows.append({
            "empresa": name,
            "periodo": period_ym,
            "conta_codigo": fr["account_code"],
            "conta_nome": ACCOUNT_NAME[fr["account_code"]],
            "valor_realizado": fr["valor_realizado"],
            "valor_orcado": fr["valor_orcado"],
        })
    return rows

def _autosize_and_style_ws(ws, header):
    """Estilo leve para a planilha (cabeçalho em negrito, largura de coluna)."""
    if not _HAS_OPENPYXL:
        return
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="0B1F3A", end_color="0B1F3A", fill_type="solid")
    for col_idx, _name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    widths = {"empresa": 26, "periodo": 12, "conta_codigo": 16, "conta_nome": 38,
              "valor_realizado": 18, "valor_orcado": 18}
    from openpyxl.utils import get_column_letter
    for col_idx, name in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(name, 16)

def write_xlsx(path: str, header: list, rows: list, sheet_title: str = "dados"):
    """Escreve um .xlsx se openpyxl existir. Retorna True se escreveu."""
    if not _HAS_OPENPYXL:
        return False
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    ws.append(header)
    for r in rows:
        if isinstance(r, dict):
            ws.append([r.get(h, "") for h in header])
        else:
            ws.append(list(r))
    _autosize_and_style_ws(ws, header)
    ws.freeze_panes = "A2"
    wb.save(path)
    return True

def write_csv_generic(path: str, header: list, rows: list):
    """Fallback CSV genérico (sempre escrito)."""
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(header)
        for r in rows:
            if isinstance(r, dict):
                w.writerow([r.get(h, "") for h in header])
            else:
                w.writerow(list(r))

def write_upload_file(base_path_noext: str, header: list, rows: list, sheet_title="dados"):
    """
    Escreve .xlsx (se possível) E sempre .csv de fallback.
    Retorna lista de caminhos efetivamente escritos.
    """
    written = []
    xlsx_path = base_path_noext + ".xlsx"
    csv_path = base_path_noext + ".csv"
    if write_xlsx(xlsx_path, header, rows, sheet_title):
        written.append(xlsx_path)
    write_csv_generic(csv_path, header, rows)
    written.append(csv_path)
    return written

def build_clean_upload_AUR_VAR(facts: list) -> list:
    """Upload limpo do mês corrente (2026-06) para AUR-VAR."""
    name = COMPANY_BY_ID["AUR-VAR"][1]
    rows = []
    for fr in facts:
        if fr["company_id"] != "AUR-VAR":
            continue
        if fr["period_date"][:7] != "2026-06":
            continue
        rows.append({
            "empresa": name,
            "periodo": "2026-06",
            "conta_codigo": fr["account_code"],
            "conta_nome": ACCOUNT_NAME[fr["account_code"]],
            "valor_realizado": fr["valor_realizado"],
            "valor_orcado": fr["valor_orcado"],
        })
    return rows

def build_invalid_upload_AUR_IND(facts: list) -> list:
    """
    Upload deliberadamente MALFORMADO para AUR-IND (2026-06), para demonstrar
    validação de schema + quarentena no pipeline de ingestão:
      * uma data inválida (periodo = '2026-13')
      * um valor não numérico ('R$ doze mil')
      * um account_code desconhecido ('CONTA_FANTASMA')
    O restante das linhas é válido para mostrar o split ok/quarentena.
    """
    name = COMPANY_BY_ID["AUR-IND"][1]
    rows = []
    base = [fr for fr in facts
            if fr["company_id"] == "AUR-IND" and fr["period_date"][:7] == "2026-06"]
    for fr in base:
        rows.append({
            "empresa": name,
            "periodo": "2026-06",
            "conta_codigo": fr["account_code"],
            "conta_nome": ACCOUNT_NAME[fr["account_code"]],
            "valor_realizado": fr["valor_realizado"],
            "valor_orcado": fr["valor_orcado"],
        })

    # --- Injeção de erros deliberados ---
    # (1) Data inválida (mês 13)
    rows.append({
        "empresa": name, "periodo": "2026-13", "conta_codigo": "R_BRUTA",
        "conta_nome": "Receita Bruta de Vendas",
        "valor_realizado": 9_999_999.99, "valor_orcado": 9_500_000.00,
    })
    # (2) Valor não numérico
    rows.append({
        "empresa": name, "periodo": "2026-06", "conta_codigo": "DESP_ADM",
        "conta_nome": "Despesas Administrativas",
        "valor_realizado": "R$ doze mil", "valor_orcado": "-450000",
    })
    # (3) account_code desconhecido
    rows.append({
        "empresa": name, "periodo": "2026-06", "conta_codigo": "CONTA_FANTASMA",
        "conta_nome": "Conta Inexistente no Plano",
        "valor_realizado": 123_456.78, "valor_orcado": 120_000.00,
    })
    return rows

# --------------------------------------------------------------------------------------
# 13. Formatação de moeda PT-BR (para textos dos insights)
# --------------------------------------------------------------------------------------
def fmt_brl_compact(v: float) -> str:
    """Formata em R$ X,Y mi / R$ X mil (pt-BR), como nas convenções do SPEC seção 10."""
    a = abs(v)
    sign = "-" if v < 0 else ""
    if a >= 1_000_000:
        s = f"{a/1_000_000:,.1f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"{sign}R$ {s} mi"
    if a >= 1_000:
        s = f"{a/1_000:,.0f}".replace(",", ".")
        return f"{sign}R$ {s} mil"
    s = f"{a:,.0f}".replace(",", ".")
    return f"{sign}R$ {s}"

def fmt_pct(v: float, casas: int = 1) -> str:
    s = f"{v:,.{casas}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{s}%"

# --------------------------------------------------------------------------------------
# 14. Helpers de série/LTM
# --------------------------------------------------------------------------------------
def last12_ending(period_list_ym: list, end_ym: str) -> list:
    """Retorna os 12 'YYYY-MM' terminando em end_ym (inclusive), em ordem ascendente."""
    if end_ym not in period_list_ym:
        end_ym = period_list_ym[-1]
    idx = period_list_ym.index(end_ym)
    start = max(0, idx - 11)
    return period_list_ym[start: idx + 1]

LTM_END_YM = ym(LAST_CLOSED)                     # SPEC: LTM termina em 2026-05
LTM_PERIODS = last12_ending(PERIOD_YM, LTM_END_YM)   # 2025-06 .. 2026-05
SPARK_PERIODS = last12_ending(PERIOD_YM, LTM_END_YM)  # 12 valores para sparklines

# --------------------------------------------------------------------------------------
# 15. Construção do dashboard_data.json (SPEC seção 8 — contrato exato)
# --------------------------------------------------------------------------------------
def safe_delta_pct(value: float, prev: float) -> float:
    if prev == 0:
        return 0.0
    return round((value / prev - 1.0) * 100.0, 2)

def build_dashboard_json(facts: list, pnl_all: dict, pos_all: dict) -> dict:
    # 15.1 Consolidado por mês (todas as derivações)
    consol = {ymk: consolidate_period(pnl_all, pos_all, ymk) for ymk in PERIOD_YM}

    # Índices úteis
    idx_last = PERIOD_YM.index(LTM_END_YM)         # 2026-05
    ym_last = LTM_END_YM
    ym_prev = PERIOD_YM[idx_last - 1]              # 2026-04

    # EBITDA LTM consolidado (soma 12 meses terminando em 2026-05)
    ebitda_ltm = sum(consol[m]["pnl"]["ebitda"] for m in LTM_PERIODS)
    receita_ltm = sum(consol[m]["pnl"]["receita_liquida"] for m in LTM_PERIODS)
    lucro_ltm = sum(consol[m]["pnl"]["lucro_liquido"] for m in LTM_PERIODS)

    # ---------------------------------------------------------------------------------
    # 15.2 meta
    # ---------------------------------------------------------------------------------
    meta = {
        "group_name": "Grupo Aurora",
        "currency": "BRL",
        "locale": "pt-BR",
        "generated_at": GENERATED_AT,
        "period_start": ym(PERIOD_START),
        "period_end": ym(PERIOD_END),
        "last_closed_period": ym(LAST_CLOSED),
        "companies": [
            {"id": cid, "name": name, "sector": sector, "color": color}
            for (cid, name, sector, color, *_rest) in COMPANIES
        ],
        "is_placeholder_brand": True,
    }

    # ---------------------------------------------------------------------------------
    # 15.3 KPIs (consolidado, último mês fechado vs anterior). spark = 12 valores.
    # ---------------------------------------------------------------------------------
    spark_receita = [round(consol[m]["pnl"]["receita_liquida"], 2) for m in SPARK_PERIODS]
    spark_ebitda = [round(consol[m]["pnl"]["ebitda"], 2) for m in SPARK_PERIODS]
    spark_lucro = [round(consol[m]["pnl"]["lucro_liquido"], 2) for m in SPARK_PERIODS]
    spark_caixa = [round(consol[m]["pos"]["caixa"], 2) for m in SPARK_PERIODS]

    caixa_now = consol[ym_last]["pos"]["caixa"]
    caixa_prev = consol[ym_prev]["pos"]["caixa"]

    receita_now = consol[ym_last]["pnl"]["receita_liquida"]
    receita_prev = consol[ym_prev]["pnl"]["receita_liquida"]
    # YoY: receita_liquida(m) / receita_liquida(m-12) - 1
    ym_year_ago = PERIOD_YM[idx_last - 12] if idx_last - 12 >= 0 else PERIOD_YM[0]
    receita_year_ago = consol[ym_year_ago]["pnl"]["receita_liquida"]
    receita_yoy = safe_delta_pct(receita_now, receita_year_ago)

    ebitda_now = consol[ym_last]["pnl"]["ebitda"]
    ebitda_prev = consol[ym_prev]["pnl"]["ebitda"]
    margem_ebitda_now = (ebitda_now / receita_now * 100.0) if receita_now else 0.0

    lucro_now = consol[ym_last]["pnl"]["lucro_liquido"]
    lucro_prev = consol[ym_prev]["pnl"]["lucro_liquido"]
    margem_liq_now = (lucro_now / receita_now * 100.0) if receita_now else 0.0

    # Dívida líquida e dívida/EBITDA (LTM)
    dl_now = consol[ym_last]["pos"]["divida_liquida"]
    dl_prev = consol[ym_prev]["pos"]["divida_liquida"]
    divida_ebitda = round(dl_now / ebitda_ltm, 2) if ebitda_ltm else 0.0

    # DSO consolidado
    dso_now = consol[ym_last]["pos"]["dso_dias"]
    dso_prev = consol[ym_prev]["pos"]["dso_dias"]

    # Capital de giro
    cg_now = consol[ym_last]["pos"]["capital_giro"]
    cg_prev = consol[ym_prev]["pos"]["capital_giro"]

    # Fluxo de caixa mensal consolidado (CAIXA(m) - CAIXA(m-1))
    fluxo_por_mes = {}
    for i, m in enumerate(PERIOD_YM):
        if i == 0:
            fluxo_por_mes[m] = 0.0
        else:
            fluxo_por_mes[m] = (consol[m]["pos"]["caixa"] -
                                consol[PERIOD_YM[i - 1]]["pos"]["caixa"])

    # Burn mensal = média( -fluxo ) dos últimos 3 meses fechados, quando negativo
    last3 = [PERIOD_YM[idx_last - 2], PERIOD_YM[idx_last - 1], ym_last]
    burns = [-fluxo_por_mes[m] for m in last3]  # positivo = queima
    burn_avg = sum(burns) / len(burns)
    if burn_avg > 0:
        runway = caixa_now / burn_avg
        runway_val = round(runway, 1)
        label_extra = f"burn {fmt_brl_compact(burn_avg)}/mês"
    else:
        runway_val = 0.0  # 0 => front interpreta como "fluxo positivo / n/a"
        label_extra = "fluxo de caixa positivo (sem burn)"

    # runway anterior (mês -1) para delta — recomputa burn 3m terminando em ym_prev
    last3_prev = [PERIOD_YM[idx_last - 3], PERIOD_YM[idx_last - 2], PERIOD_YM[idx_last - 1]]
    burns_prev = [-fluxo_por_mes[m] for m in last3_prev]
    burn_avg_prev = sum(burns_prev) / len(burns_prev)
    if burn_avg_prev > 0:
        runway_prev_val = round(caixa_prev / burn_avg_prev, 1)
    else:
        runway_prev_val = 0.0

    kpis = {
        "caixa": {
            "value": round(caixa_now, 2),
            "prev": round(caixa_prev, 2),
            "delta_pct": safe_delta_pct(caixa_now, caixa_prev),
            "spark": spark_caixa,
        },
        "runway_meses": {
            "value": runway_val,
            "prev": runway_prev_val,
            "delta_pct": safe_delta_pct(runway_val, runway_prev_val),
            "label_extra": label_extra,
        },
        "receita_liquida": {
            "value": round(receita_now, 2),
            "prev": round(receita_prev, 2),
            "delta_pct": safe_delta_pct(receita_now, receita_prev),
            "spark": spark_receita,
            "yoy_pct": receita_yoy,
        },
        "ebitda": {
            "value": round(ebitda_now, 2),
            "prev": round(ebitda_prev, 2),
            "delta_pct": safe_delta_pct(ebitda_now, ebitda_prev),
            "spark": spark_ebitda,
            "margem_pct": round(margem_ebitda_now, 2),
        },
        "lucro_liquido": {
            "value": round(lucro_now, 2),
            "prev": round(lucro_prev, 2),
            "delta_pct": safe_delta_pct(lucro_now, lucro_prev),
            "spark": spark_lucro,
            "margem_pct": round(margem_liq_now, 2),
        },
        "divida_liquida": {
            "value": round(dl_now, 2),
            "prev": round(dl_prev, 2),
            "delta_pct": safe_delta_pct(dl_now, dl_prev),
            "divida_ebitda": divida_ebitda,
        },
        "dso_dias": {
            "value": round(dso_now, 1),
            "prev": round(dso_prev, 1),
            "delta_pct": safe_delta_pct(dso_now, dso_prev),
        },
        "capital_giro": {
            "value": round(cg_now, 2),
            "prev": round(cg_prev, 2),
            "delta_pct": safe_delta_pct(cg_now, cg_prev),
        },
    }

    # ---------------------------------------------------------------------------------
    # 15.4 series_mensal (consolidado, ascendente por "YYYY-MM")
    # ---------------------------------------------------------------------------------
    # Orçado consolidado por mês (receita líquida orçada e EBITDA orçado)
    def consolidated_budget_for_period(ymk: str) -> dict:
        """Soma orçado das contas-folha de todas as empresas e deriva receita/ebitda orçados."""
        summed = {code: 0.0 for code in PNL_CODES}
        for fr in facts:
            if fr["period_date"][:7] != ymk:
                continue
            if fr["account_code"] in summed:
                summed[fr["account_code"]] += fr["valor_orcado"]
        d = derive_pnl(summed)
        return {"receita_orcada": d["receita_liquida"], "ebitda_orcado": d["ebitda"]}

    series_mensal = []
    for ymk in PERIOD_YM:
        c = consol[ymk]
        budget = consolidated_budget_for_period(ymk)
        series_mensal.append({
            "period": ymk,
            "receita_bruta": round(c["pnl"]["receita_bruta"], 2),
            "receita_liquida": round(c["pnl"]["receita_liquida"], 2),
            "lucro_bruto": round(c["pnl"]["lucro_bruto"], 2),
            "ebitda": round(c["pnl"]["ebitda"], 2),
            "ebit": round(c["pnl"]["ebit"], 2),
            "lucro_liquido": round(c["pnl"]["lucro_liquido"], 2),
            "caixa": round(c["pos"]["caixa"], 2),
            "divida_liquida": round(c["pos"]["divida_liquida"], 2),
            "fluxo_caixa": round(fluxo_por_mes[ymk], 2),
            "margem_ebitda_pct": round(c["pnl"]["margem_ebitda_pct"], 2),
            "receita_orcada": round(budget["receita_orcada"], 2),
            "ebitda_orcado": round(budget["ebitda_orcado"], 2),
        })

    # ---------------------------------------------------------------------------------
    # 15.5 por_empresa (LTM + sparklines de receita/ebitda + share + YoY)
    # ---------------------------------------------------------------------------------
    por_empresa = []
    # receita líquida LTM consolidada para share (exclui ELIM ~0 não distorce)
    receita_ltm_por_emp = {}
    for (cid, name, sector, color, *_rest) in COMPANIES:
        # Derivar por empresa-mês
        emp_pnl_month = {}
        for ymk in PERIOD_YM:
            emp_pnl_month[ymk] = derive_pnl(pnl_all[cid][ymk])
        rl_ltm = sum(emp_pnl_month[m]["receita_liquida"] for m in LTM_PERIODS)
        eb_ltm = sum(emp_pnl_month[m]["ebitda"] for m in LTM_PERIODS)
        ll_ltm = sum(emp_pnl_month[m]["lucro_liquido"] for m in LTM_PERIODS)
        margem_eb = (eb_ltm / rl_ltm * 100.0) if rl_ltm else 0.0
        # YoY por empresa: receita líquida último mês vs 12 meses antes
        rl_last = emp_pnl_month[ym_last]["receita_liquida"]
        rl_year_ago = emp_pnl_month[ym_year_ago]["receita_liquida"]
        yoy = safe_delta_pct(rl_last, rl_year_ago)
        serie_receita = [round(emp_pnl_month[m]["receita_liquida"], 2) for m in SPARK_PERIODS]
        serie_ebitda = [round(emp_pnl_month[m]["ebitda"], 2) for m in SPARK_PERIODS]
        receita_ltm_por_emp[cid] = rl_ltm
        por_empresa.append({
            "company_id": cid,
            "name": name,
            "sector": sector,
            "color": color,
            "receita_ltm": round(rl_ltm, 2),
            "ebitda_ltm": round(eb_ltm, 2),
            "margem_ebitda_pct": round(margem_eb, 2),
            "lucro_liquido_ltm": round(ll_ltm, 2),
            "share_receita_pct": 0.0,  # preenchido abaixo
            "yoy_pct": yoy,
            "serie_receita": serie_receita,
            "serie_ebitda": serie_ebitda,
        })
    total_rl_ltm = sum(max(v, 0.0) for v in receita_ltm_por_emp.values())
    for row in por_empresa:
        rl = receita_ltm_por_emp[row["company_id"]]
        row["share_receita_pct"] = round((rl / total_rl_ltm * 100.0), 2) if total_rl_ltm else 0.0

    # ---------------------------------------------------------------------------------
    # 15.6 dre_consolidada (último mês fechado + LTM + orçado do mês + var %)
    # ---------------------------------------------------------------------------------
    # Orçado consolidado do último mês (por conta-folha) para derivar linhas orçadas
    summed_budget_last = {code: 0.0 for code in PNL_CODES}
    for fr in facts:
        if fr["period_date"][:7] == ym_last and fr["account_code"] in summed_budget_last:
            summed_budget_last[fr["account_code"]] += fr["valor_orcado"]
    dre_budget = derive_pnl(summed_budget_last)

    c_last = consol[ym_last]["pnl"]

    def ltm_sum(key: str) -> float:
        return sum(consol[m]["pnl"][key] for m in LTM_PERIODS)

    def dre_line(linha, code, mes_val, ltm_val, orcado_val):
        var = safe_delta_pct(mes_val, orcado_val)
        return {"linha": linha, "code": code,
                "mes": round(mes_val, 2), "ltm": round(ltm_val, 2),
                "orcado_mes": round(orcado_val, 2), "var_pct": var}

    # Linhas-folha relevantes consolidadas do último mês (realizado) e orçado
    leaf_last = consol[ym_last]["pnl_leaf"]
    dre_consolidada = [
        dre_line("Receita Líquida", "receita_liquida",
                 c_last["receita_liquida"], ltm_sum("receita_liquida"), dre_budget["receita_liquida"]),
        dre_line("(-) CMV", "CMV",
                 leaf_last["CMV"], ltm_sum_leaf := sum(consol[m]["pnl_leaf"]["CMV"] for m in LTM_PERIODS),
                 summed_budget_last["CMV"]),
        dre_line("Lucro Bruto", "lucro_bruto",
                 c_last["lucro_bruto"], ltm_sum("lucro_bruto"), dre_budget["lucro_bruto"]),
        dre_line("(-) Despesas com Pessoal", "DESP_PESSOAL",
                 leaf_last["DESP_PESSOAL"], sum(consol[m]["pnl_leaf"]["DESP_PESSOAL"] for m in LTM_PERIODS),
                 summed_budget_last["DESP_PESSOAL"]),
        dre_line("(-) Despesas Comerciais", "DESP_VENDAS",
                 leaf_last["DESP_VENDAS"], sum(consol[m]["pnl_leaf"]["DESP_VENDAS"] for m in LTM_PERIODS),
                 summed_budget_last["DESP_VENDAS"]),
        dre_line("(-) Despesas Administrativas", "DESP_ADM",
                 leaf_last["DESP_ADM"], sum(consol[m]["pnl_leaf"]["DESP_ADM"] for m in LTM_PERIODS),
                 summed_budget_last["DESP_ADM"]),
        dre_line("(-) Outras Despesas Operacionais", "DESP_OUTRAS",
                 leaf_last["DESP_OUTRAS"], sum(consol[m]["pnl_leaf"]["DESP_OUTRAS"] for m in LTM_PERIODS),
                 summed_budget_last["DESP_OUTRAS"]),
        dre_line("EBITDA", "ebitda",
                 c_last["ebitda"], ltm_sum("ebitda"), dre_budget["ebitda"]),
        dre_line("(-) Depreciação e Amortização", "DEPRECIACAO",
                 leaf_last["DEPRECIACAO"], sum(consol[m]["pnl_leaf"]["DEPRECIACAO"] for m in LTM_PERIODS),
                 summed_budget_last["DEPRECIACAO"]),
        dre_line("EBIT", "ebit",
                 c_last["ebit"], ltm_sum("ebit"), dre_budget["ebit"]),
        dre_line("Resultado Financeiro", "RESULT_FIN",
                 leaf_last["RESULT_FIN"], sum(consol[m]["pnl_leaf"]["RESULT_FIN"] for m in LTM_PERIODS),
                 summed_budget_last["RESULT_FIN"]),
        dre_line("(-) IR e CSLL", "IRPJ_CSLL",
                 leaf_last["IRPJ_CSLL"], sum(consol[m]["pnl_leaf"]["IRPJ_CSLL"] for m in LTM_PERIODS),
                 summed_budget_last["IRPJ_CSLL"]),
        dre_line("Lucro Líquido", "lucro_liquido",
                 c_last["lucro_liquido"], ltm_sum("lucro_liquido"), dre_budget["lucro_liquido"]),
    ]

    # ---------------------------------------------------------------------------------
    # 15.7 gastos_por_categoria (donut) — OPEX + CMV, último mês, consolidado
    # ---------------------------------------------------------------------------------
    cat_defs = [
        ("CMV",          "CMV / Custo dos Serviços", "#4F6BED"),
        ("DESP_PESSOAL", "Pessoal",                   "#0EA5E9"),
        ("DESP_VENDAS",  "Comercial e Marketing",     "#10B981"),
        ("DESP_ADM",     "Administrativas",           "#F59E0B"),
        ("DESP_OUTRAS",  "Outras Operacionais",       "#6B7280"),
        ("DEPRECIACAO",  "Depreciação",               "#94A3B8"),
    ]
    gastos_vals = []
    for code, label, color in cat_defs:
        val = abs(leaf_last[code])
        gastos_vals.append((label, val, color))
    total_gastos = sum(v for _, v, _ in gastos_vals)
    gastos_por_categoria = []
    for label, val, color in gastos_vals:
        pct = (val / total_gastos * 100.0) if total_gastos else 0.0
        gastos_por_categoria.append({
            "categoria": label,
            "valor": round(val, 2),
            "pct": round(pct, 2),
            "color": color,
        })

    # ---------------------------------------------------------------------------------
    # 15.8 orcado_vs_realizado (por empresa, último mês — barras de variância)
    # ---------------------------------------------------------------------------------
    # Usamos receita líquida realizada vs orçada por empresa no último mês fechado.
    orcado_vs_realizado = []
    for (cid, name, sector, color, *_rest) in COMPANIES:
        real_leaf = pnl_all[cid][ym_last]
        realizado = real_leaf["R_BRUTA"] + real_leaf["DEDUCOES"]
        # orçado da empresa (soma das folhas de receita)
        orc_rb = 0.0
        orc_ded = 0.0
        for fr in facts:
            if fr["company_id"] == cid and fr["period_date"][:7] == ym_last:
                if fr["account_code"] == "R_BRUTA":
                    orc_rb = fr["valor_orcado"]
                elif fr["account_code"] == "DEDUCOES":
                    orc_ded = fr["valor_orcado"]
        orcado = orc_rb + orc_ded
        var = safe_delta_pct(realizado, orcado)
        orcado_vs_realizado.append({
            "company_id": cid,
            "name": name,
            "realizado": round(realizado, 2),
            "orcado": round(orcado, 2),
            "var_pct": var,
        })

    # ---------------------------------------------------------------------------------
    # 15.9 aging (AR/AP, consolidado, último mês) — buckets sintéticos coerentes
    # ---------------------------------------------------------------------------------
    ar_total = consol[ym_last]["pos"]["ar"]
    ap_total = consol[ym_last]["pos"]["ap"]
    # Distribuição típica: maior parte corrente (0-30), cauda crescente em atraso.
    ar_dist = [0.62, 0.22, 0.10, 0.06]   # 0-30,31-60,61-90,90+
    ap_dist = [0.70, 0.18, 0.08, 0.04]
    faixas = ["0-30", "31-60", "61-90", "90+"]
    aging = {
        "receber": [{"faixa": faixas[i], "valor": round(ar_total * ar_dist[i], 2)}
                    for i in range(4)],
        "pagar": [{"faixa": faixas[i], "valor": round(ap_total * ap_dist[i], 2)}
                  for i in range(4)],
    }

    # ---------------------------------------------------------------------------------
    # 15.10 insights_ia (comentário de variância em PT-BR, ancorado em números reais)
    # ---------------------------------------------------------------------------------
    insights_ia = build_insights(
        consol=consol, ym_last=ym_last, ym_prev=ym_prev, ym_year_ago=ym_year_ago,
        por_empresa=por_empresa, orcado_vs_realizado=orcado_vs_realizado,
        dre_consolidada=dre_consolidada, kpis=kpis, ebitda_ltm=ebitda_ltm,
        receita_yoy=receita_yoy, fluxo_por_mes=fluxo_por_mes, facts=facts,
        pnl_all=pnl_all, divida_ebitda=divida_ebitda,
    )

    # ---------------------------------------------------------------------------------
    # 15.11 perguntas_sugeridas
    # ---------------------------------------------------------------------------------
    perguntas_sugeridas = [
        "Qual o EBITDA consolidado do último mês fechado?",
        "Como está a margem EBITDA da Aurora Varejo nos últimos 12 meses?",
        "Quais empresas estouraram o orçamento de despesas no mês?",
        "Qual a dívida líquida consolidada e a alavancagem (dívida/EBITDA)?",
        "Qual o runway de caixa do grupo no ritmo atual?",
        "Como evoluiu a receita líquida consolidada ano contra ano?",
        "Qual empresa tem a maior participação na receita do grupo?",
    ]

    # ---------------------------------------------------------------------------------
    # 15.12 respostas_demo (respostas canônicas ancoradas, demo offline)
    # ---------------------------------------------------------------------------------
    respostas_demo = build_respostas_demo(
        consol=consol, ym_last=ym_last, ebitda_ltm=ebitda_ltm, receita_ltm=receita_ltm,
        por_empresa=por_empresa, kpis=kpis, divida_ebitda=divida_ebitda,
        receita_yoy=receita_yoy, orcado_vs_realizado=orcado_vs_realizado,
    )

    # ---------------------------------------------------------------------------------
    # 15.13 montagem final
    # ---------------------------------------------------------------------------------
    return {
        "meta": meta,
        "kpis": kpis,
        "series_mensal": series_mensal,
        "por_empresa": por_empresa,
        "dre_consolidada": dre_consolidada,
        "gastos_por_categoria": gastos_por_categoria,
        "orcado_vs_realizado": orcado_vs_realizado,
        "aging": aging,
        "insights_ia": insights_ia,
        "perguntas_sugeridas": perguntas_sugeridas,
        "respostas_demo": respostas_demo,
    }

# --------------------------------------------------------------------------------------
# 16. Insights de IA (grounded) e respostas demo
# --------------------------------------------------------------------------------------
def build_insights(*, consol, ym_last, ym_prev, ym_year_ago, por_empresa,
                   orcado_vs_realizado, dre_consolidada, kpis, ebitda_ltm,
                   receita_yoy, fluxo_por_mes, facts, pnl_all, divida_ebitda):
    insights = []

    # (a) Receita consolidada MoM
    rl_now = consol[ym_last]["pnl"]["receita_liquida"]
    rl_prev = consol[ym_prev]["pnl"]["receita_liquida"]
    mom = safe_delta_pct(rl_now, rl_prev)
    sev = "positive" if mom >= 0 else "warning"
    insights.append({
        "severity": sev,
        "titulo": f"Receita líquida consolidada {'cresceu' if mom >= 0 else 'recuou'} {fmt_pct(abs(mom))} no mês",
        "texto": (f"Em {ym_last}, a receita líquida consolidada foi de {fmt_brl_compact(rl_now)}, "
                  f"ante {fmt_brl_compact(rl_prev)} em {ym_prev} "
                  f"({'+' if mom >= 0 else ''}{fmt_pct(mom)} no comparativo mensal). "
                  f"No acumulado de 12 meses, a variação ano contra ano é de {fmt_pct(receita_yoy)}."),
    })

    # (b) EBITDA e margem
    eb_now = consol[ym_last]["pnl"]["ebitda"]
    me_now = consol[ym_last]["pnl"]["margem_ebitda_pct"]
    insights.append({
        "severity": "info",
        "titulo": f"EBITDA do mês em {fmt_brl_compact(eb_now)} (margem {fmt_pct(me_now)})",
        "texto": (f"O EBITDA consolidado de {ym_last} atingiu {fmt_brl_compact(eb_now)}, "
                  f"com margem EBITDA de {fmt_pct(me_now)} sobre a receita líquida. "
                  f"No acumulado LTM (12 meses até {ym_last}), o EBITDA soma {fmt_brl_compact(ebitda_ltm)}."),
    })

    # (c) Maior estouro de orçamento de receita (pior var negativa)
    pior = min(orcado_vs_realizado, key=lambda r: r["var_pct"])
    melhor = max(orcado_vs_realizado, key=lambda r: r["var_pct"])
    if pior["var_pct"] < 0:
        insights.append({
            "severity": "warning",
            "titulo": f"{pior['name']} abaixo do orçado de receita em {fmt_pct(abs(pior['var_pct']))}",
            "texto": (f"A {pior['name']} realizou {fmt_brl_compact(pior['realizado'])} de receita líquida "
                      f"em {ym_last}, contra orçamento de {fmt_brl_compact(pior['orcado'])} "
                      f"({fmt_pct(pior['var_pct'])} vs. orçado). Recomenda-se revisar pipeline comercial e mix."),
        })

    # (d) Melhor performance vs orçado
    if melhor["var_pct"] > 0:
        insights.append({
            "severity": "positive",
            "titulo": f"{melhor['name']} superou o orçamento de receita em {fmt_pct(melhor['var_pct'])}",
            "texto": (f"A {melhor['name']} entregou {fmt_brl_compact(melhor['realizado'])} de receita líquida "
                      f"em {ym_last}, acima dos {fmt_brl_compact(melhor['orcado'])} orçados "
                      f"(+{fmt_pct(melhor['var_pct'])}). Destaque positivo do período."),
        })

    # (e) Alavancagem (dívida líquida / EBITDA)
    dl = kpis["divida_liquida"]["value"]
    sev_lev = "warning" if divida_ebitda > 2.5 else "info"
    insights.append({
        "severity": sev_lev,
        "titulo": f"Alavancagem em {divida_ebitda:.2f}x dívida líquida/EBITDA",
        "texto": (f"A dívida líquida consolidada encerrou {ym_last} em {fmt_brl_compact(dl)}, "
                  f"equivalente a {divida_ebitda:.2f}x o EBITDA LTM. "
                  f"{'Patamar exige atenção ao serviço da dívida.' if divida_ebitda > 2.5 else 'Patamar considerado confortável.'}"),
    })

    # (f) Empresa líder em receita (share)
    lider = max((r for r in por_empresa if r["company_id"] != "ELIM"),
                key=lambda r: r["share_receita_pct"])
    insights.append({
        "severity": "info",
        "titulo": f"{lider['name']} concentra {fmt_pct(lider['share_receita_pct'])} da receita LTM",
        "texto": (f"A {lider['name']} ({lider['sector']}) responde por {fmt_pct(lider['share_receita_pct'])} "
                  f"da receita líquida consolidada nos últimos 12 meses ({fmt_brl_compact(lider['receita_ltm'])}), "
                  f"com margem EBITDA de {fmt_pct(lider['margem_ebitda_pct'])} e crescimento YoY de {fmt_pct(lider['yoy_pct'])}."),
    })

    # Limitar a 4-6 entradas (SPEC). Mantém as 6 mais relevantes na ordem construída.
    return insights[:6]

def build_respostas_demo(*, consol, ym_last, ebitda_ltm, receita_ltm, por_empresa,
                         kpis, divida_ebitda, receita_yoy, orcado_vs_realizado):
    eb_now = consol[ym_last]["pnl"]["ebitda"]
    me_now = consol[ym_last]["pnl"]["margem_ebitda_pct"]
    rl_now = consol[ym_last]["pnl"]["receita_liquida"]
    dl = kpis["divida_liquida"]["value"]
    runway = kpis["runway_meses"]["value"]

    varejo = next(r for r in por_empresa if r["company_id"] == "AUR-VAR")
    lider = max((r for r in por_empresa if r["company_id"] != "ELIM"),
                key=lambda r: r["share_receita_pct"])
    pior = min(orcado_vs_realizado, key=lambda r: r["var_pct"])

    respostas = [
        {
            "q": "Qual o EBITDA consolidado do último mês fechado?",
            "a": (f"No último mês fechado ({ym_last}), o EBITDA consolidado do Grupo Aurora foi de "
                  f"{fmt_brl_compact(eb_now)}, com margem EBITDA de {fmt_pct(me_now)} sobre a receita "
                  f"líquida de {fmt_brl_compact(rl_now)}. No acumulado LTM, o EBITDA soma "
                  f"{fmt_brl_compact(ebitda_ltm)}."),
            "fontes": [f"DRE consolidada {ym_last}", "v_pnl_consolidado_month", "v_kpi_consolidado_ltm"],
        },
        {
            "q": "Como está a margem EBITDA da Aurora Varejo nos últimos 12 meses?",
            "a": (f"A Aurora Varejo S.A. apresentou receita líquida LTM de {fmt_brl_compact(varejo['receita_ltm'])} "
                  f"e EBITDA LTM de {fmt_brl_compact(varejo['ebitda_ltm'])}, o que representa margem EBITDA de "
                  f"{fmt_pct(varejo['margem_ebitda_pct'])}. O crescimento de receita ano contra ano foi de "
                  f"{fmt_pct(varejo['yoy_pct'])}."),
            "fontes": ["v_pnl_company_month (AUR-VAR)", f"DRE AUR-VAR LTM até {ym_last}"],
        },
        {
            "q": "Quais empresas estouraram o orçamento de despesas no mês?",
            "a": (f"No comparativo de receita realizada vs. orçada de {ym_last}, o maior desvio negativo foi da "
                  f"{pior['name']}: realizou {fmt_brl_compact(pior['realizado'])} contra "
                  f"{fmt_brl_compact(pior['orcado'])} orçados ({fmt_pct(pior['var_pct'])}). "
                  f"As variações deliberadas de orçamento também sinalizam pressão em despesas comerciais "
                  f"(Aurora Varejo) e CMV (Aurora Indústria)."),
            "fontes": ["v_budget_vs_actual", f"Orçado vs Realizado {ym_last}"],
        },
        {
            "q": "Qual a dívida líquida consolidada e a alavancagem (dívida/EBITDA)?",
            "a": (f"A dívida líquida consolidada em {ym_last} é de {fmt_brl_compact(dl)}, equivalente a "
                  f"{divida_ebitda:.2f}x o EBITDA LTM ({fmt_brl_compact(ebitda_ltm)}). "
                  f"{'É um patamar que requer monitoramento do serviço da dívida.' if divida_ebitda > 2.5 else 'É um patamar de alavancagem confortável.'}"),
            "fontes": ["v_position_company_month", "v_kpi_consolidado_ltm"],
        },
        {
            "q": "Qual o runway de caixa do grupo no ritmo atual?",
            "a": (f"O caixa consolidado encerrou {ym_last} em {fmt_brl_compact(kpis['caixa']['value'])}. "
                  + (f"Considerando a queima média dos últimos 3 meses, o runway estimado é de "
                     f"{runway:.1f} meses ({kpis['runway_meses']['label_extra']})."
                     if runway and runway > 0 else
                     "O grupo apresenta fluxo de caixa operacional positivo no período, sem queima de caixa (runway não aplicável).")),
            "fontes": ["series_mensal (fluxo_caixa)", "v_position_company_month"],
        },
        {
            "q": "Qual empresa tem a maior participação na receita do grupo?",
            "a": (f"A {lider['name']} ({lider['sector']}) lidera com {fmt_pct(lider['share_receita_pct'])} de "
                  f"participação na receita líquida consolidada LTM ({fmt_brl_compact(lider['receita_ltm'])}), "
                  f"seguida pelas demais unidades de negócio do Grupo Aurora."),
            "fontes": ["v_pnl_company_month", "por_empresa (share_receita_pct)"],
        },
    ]
    return respostas

# --------------------------------------------------------------------------------------
# 17. Escrita do JSON (out + cópia para dashboard)
# --------------------------------------------------------------------------------------
def write_json(payload: dict):
    os.makedirs(OUT_DIR, exist_ok=True)
    os.makedirs(DASHBOARD_DIR, exist_ok=True)
    with open(DASHBOARD_JSON_OUT, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    with open(DASHBOARD_JSON_COPY, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

# --------------------------------------------------------------------------------------
# 18. Main
# --------------------------------------------------------------------------------------
def main():
    # Garante diretórios
    for d in (OUT_DIR, RAW_DIR, DASHBOARD_DIR):
        os.makedirs(d, exist_ok=True)

    # 18.1 Gera fatos + estruturas
    facts, pnl_all, pos_all = build_all_facts()

    # 18.2 Escreve a tabela fato longa
    n_facts = write_fact_csv(facts)

    # 18.3 Planilhas históricas por empresa (xlsx + csv fallback)
    raw_written = []
    for (cid, *_rest) in COMPANIES:
        rows = upload_rows_for_company(cid, facts)
        base = os.path.join(RAW_DIR, f"historico_{cid}")
        raw_written += write_upload_file(base, UPLOAD_COLS, rows, sheet_title=cid)

    # 18.4 Upload limpo do mês corrente (AUR-VAR 2026-06)
    clean_rows = build_clean_upload_AUR_VAR(facts)
    base_clean = os.path.join(RAW_DIR, "upload_AUR-VAR_2026-06")
    raw_written += write_upload_file(base_clean, UPLOAD_COLS, clean_rows, sheet_title="AUR-VAR 2026-06")

    # 18.5 Upload INVÁLIDO (AUR-IND 2026-06) — validação/quarentena
    invalid_rows = build_invalid_upload_AUR_IND(facts)
    base_invalid = os.path.join(RAW_DIR, "upload_AUR-IND_2026-06_INVALIDO")
    raw_written += write_upload_file(base_invalid, UPLOAD_COLS, invalid_rows,
                                     sheet_title="AUR-IND INVALIDO")

    # 18.6 dashboard_data.json (contrato) + cópia
    payload = build_dashboard_json(facts, pnl_all, pos_all)
    write_json(payload)

    # 18.7 Resumo consolidado (para o print final)
    ltm_periods = LTM_PERIODS
    consol = {ymk: consolidate_period(pnl_all, pos_all, ymk) for ymk in PERIOD_YM}
    receita_ltm = sum(consol[m]["pnl"]["receita_liquida"] for m in ltm_periods)
    ebitda_ltm = sum(consol[m]["pnl"]["ebitda"] for m in ltm_periods)

    # ----------------------------------------------------------------------------------
    # 18.8 Sumário no console
    # ----------------------------------------------------------------------------------
    n_companies = len(COMPANIES)
    n_periods = len(PERIODS)
    n_xlsx = sum(1 for p in raw_written if p.endswith(".xlsx"))
    n_csv_raw = sum(1 for p in raw_written if p.endswith(".csv"))

    print("=" * 70)
    print("Cockpit Financeiro Estratégico — geração de dados (determinística, seed=42)")
    print("=" * 70)
    if not _HAS_OPENPYXL:
        print("[nota] openpyxl não encontrado: gerados apenas arquivos .csv de fallback")
        print("       (instale 'openpyxl' para também produzir os .xlsx de ingestão).")
    print(f"Empresas .................. {n_companies}  (perímetro de consolidação)")
    print(f"Períodos .................. {n_periods}  ({PERIOD_YM[0]} .. {PERIOD_YM[-1]})")
    print(f"Último mês fechado ........ {ym(LAST_CLOSED)}  | LTM: {ltm_periods[0]} .. {ltm_periods[-1]}")
    print(f"Linhas em fact_financials . {n_facts}")
    print(f"  -> {FACT_CSV}")
    print(f"Planilhas de ingestão ..... {n_xlsx} .xlsx + {n_csv_raw} .csv (em data/raw/)")
    print(f"  -> historico_<empresa>, upload_AUR-VAR_2026-06, upload_AUR-IND_2026-06_INVALIDO")
    print(f"dashboard_data.json ....... escrito em:")
    print(f"  -> {DASHBOARD_JSON_OUT}")
    print(f"  -> {DASHBOARD_JSON_COPY}")
    print("-" * 70)
    print(f"Consolidado LTM (12 meses até {ym(LAST_CLOSED)}):")
    print(f"  receita_liquida_ltm ..... {fmt_brl_compact(receita_ltm)}  (R$ {receita_ltm:,.2f})")
    print(f"  ebitda_ltm .............. {fmt_brl_compact(ebitda_ltm)}  (R$ {ebitda_ltm:,.2f})")
    margem = (ebitda_ltm / receita_ltm * 100.0) if receita_ltm else 0.0
    print(f"  margem_ebitda_ltm ....... {fmt_pct(margem)}")
    print("=" * 70)
    print("OK — dados regenerados com sucesso.")


if __name__ == "__main__":
    main()
