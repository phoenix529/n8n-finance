#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
adapter_dre.py — adaptador das planilhas reais "DRE Acumulado" (Ref Comunicação).

Cada arquivo (uma unidade: Zup, Viv, REF+, BD, 4PR) tem uma aba **DRE-Base** com:
  - linhas contábeis na coluna B (com código/descrição em C/D)
  - meses como colunas DATADAS (E..P = jan..dez 2026)
Este adaptador lê essa aba, casa cada linha de interesse por TEXTO (robusto a
posição), e produz fatos canônicos:  (unidade, periodo 'YYYY-MM', metrica, valor).

Métricas extraídas (o arquivo já traz os subtotais calculados — confiamos neles):
  receita_bruta, receita_liquida, resultado_agencia (lucro bruto), ebit,
  resultado_liquido, geracao_caixa
Categorias de gasto (magnitude, para "gastos por categoria"):
  deducoes, custos, desp_pessoal, desp_infra, desp_outras, desp_adm, tributos

Uso:
  python data/adapter_dre.py                 # relatório + reconciliação (todas as unidades)
  from adapter_dre import extract_file        # -> [{unidade, periodo, metrica, valor}, ...]
"""
import os, glob, json, unicodedata, datetime as dt
import openpyxl

UNITS = {  # arquivo (prefixo) -> (codigo_unidade, nome)
    "Zup": ("ZUP", "Zup"), "Viv": ("VIV", "Viv"), "REF+": ("REFMAIS", "REF+"),
    "BD": ("BD", "BD"), "4PR": ("4PR", "4PR"),
}

# Janela de ATUAIS: hoje é 2026-06; meses posteriores nas planilhas são projeção
# (cauda repetida/alternada Jul–Dez). Tudo após o cutoff é descartado da ingestão.
MAX_PERIOD = os.environ.get("DRE_MAX_PERIOD", "2026-06")

# (codigo, rótulo, tipo, [(padrão_normalizado, modo)])
#   tipo: 'metric' (subtotal calculado, sinal do arquivo) | 'cat' (gasto, magnitude)
#   modo: 'eq' (igualdade exata normalizada) | 'pre' (startswith)
# Cada arquivo rotula as linhas de forma ligeiramente diferente (AGENCIA/PRODUTORA/
# EMPRESA; "RESULTADO LIQUIDO" vs "RESULTADO ANTES DAS PARTICIPAÇÕES"): por isso
# múltiplos padrões por linha. A 1ª linha que casa (de cima p/ baixo) vence.
LINES = [
    ("receita_bruta",     "Receita Bruta",                      "metric", [("receita bruta", "eq")]),
    ("deducoes",          "Deduções e Impostos",                "cat",    [("deducoes", "pre")]),
    ("receita_liquida",   "Receita Operacional Líquida",        "metric", [("receita operacional liquida", "eq")]),
    ("custos",            "Custos dos Serviços Vendidos",       "cat",    [("custos dos servicos", "pre")]),
    ("resultado_agencia", "Resultado Operacional (Lucro Bruto)","metric", [("resultado op da agencia", "pre"),
                                                                            ("resultado op agencia", "pre"),
                                                                            ("resultado op da produtora", "pre"),
                                                                            ("resultado op da empresa", "pre")]),
    ("desp_pessoal",      "Gastos com Pessoal",                 "cat",    [("gastos com pessoal", "eq")]),
    ("desp_infra",        "Infraestrutura",                     "cat",    [("infra estrutura", "pre"), ("infraestrutura", "pre")]),
    ("desp_outras",       "Outras Despesas",                    "cat",    [("outras despesas", "eq")]),
    ("desp_adm",          "Despesas Administrativas",           "cat",    [("despesas adm", "pre")]),
    ("tributos",          "Tributos Federais (IRPJ+CSLL)",      "cat",    [("tributos federais", "pre")]),
    ("ebit",              "Resultado Operacional (EBIT)",       "metric", [("resultado operacional antes dos impostos", "eq")]),
    ("resultado_liquido", "Resultado Líquido",                  "metric", [("resultado antes das participacoes", "pre"),
                                                                            ("resultado liquido", "eq")]),
    ("geracao_caixa",     "Geração de Caixa",                   "metric", [("geracao de caixa", "pre")]),
]


def norm(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    s = "".join(c.lower() if c.isalnum() else " " for c in s)
    return " ".join(s.split())


def find_dre_sheet(wb):
    for nm in wb.sheetnames:
        if norm(nm).replace(" ", "") in ("drebase", "dre"):
            return wb[nm]
    for nm in wb.sheetnames:
        if "dre" in norm(nm):
            return wb[nm]
    return None


def find_header(ws):
    """Localiza a linha de cabeçalho com as datas mensais; retorna (row, {col: 'YYYY-MM'})."""
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=12, max_col=22, values_only=True), 1):
        months = {}
        for ci, v in enumerate(row, 1):
            if isinstance(v, dt.datetime) and v.year == 2026:
                months[ci] = f"{v.year}-{v.month:02d}"
        if len(months) >= 6:
            return r, months
    return None, {}


def _match(text, pairs):
    n = norm(text)
    if not n:
        return False
    for p, mode in pairs:
        if mode == "eq" and n == p:
            return True
        if mode == "pre" and n.startswith(p):
            return True
    return False


def extract_file(path, cutoff=MAX_PERIOD):
    """Retorna (facts, ucode, uname, periods_atuais).
    facts = [{unidade, unidade_nome, periodo, code, label, tipo, valor}] — só meses <= cutoff."""
    base = os.path.basename(path)
    prefix = base.split(" - ")[0].strip()
    ucode, uname = UNITS.get(prefix, (norm(prefix).upper().replace(" ", "_"), prefix))
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = find_dre_sheet(wb)
    if ws is None:
        raise RuntimeError(f"{base}: aba DRE-Base não encontrada (abas: {wb.sheetnames[:6]}...)")
    hrow, months = find_header(ws)
    if not months:
        raise RuntimeError(f"{base}: cabeçalho de meses (datas 2026) não encontrado")
    months = {ci: per for ci, per in months.items() if per <= cutoff}   # só ATUAIS
    facts = []
    taken = set()
    for row in ws.iter_rows(min_row=hrow + 1, max_row=ws.max_row, values_only=True):
        label_cell = row[1] if len(row) > 1 else None   # coluna B
        if label_cell is None:
            continue
        for code, lbl, tipo, pairs in LINES:
            if code in taken:
                continue
            if _match(label_cell, pairs):
                taken.add(code)
                for ci, per in months.items():
                    v = row[ci - 1] if ci - 1 < len(row) else None
                    if isinstance(v, (int, float)):
                        facts.append({"unidade": ucode, "unidade_nome": uname, "periodo": per,
                                      "code": code, "label": lbl, "tipo": tipo, "valor": float(v)})
                break
    wb.close()
    return facts, ucode, uname, sorted(months.values())


def build_all_facts(cutoff=MAX_PERIOD, incoming=None):
    """Roda o adaptador em todas as 5 planilhas e devolve a lista única de fatos."""
    incoming = incoming or os.path.join(os.path.dirname(__file__), "incoming")
    out = []
    for f in sorted(glob.glob(os.path.join(incoming, "*DRE*.xlsx"))):
        facts, *_ = extract_file(f, cutoff=cutoff)
        out.extend(facts)
    return out


def _fmt(v):
    return f"R$ {v/1e6:.2f} mi" if abs(v) >= 1e6 else f"R$ {v/1e3:.0f} mil"


def main():
    files = sorted(glob.glob(os.path.join(os.path.dirname(__file__), "incoming", "*DRE*.xlsx")))
    if not files:
        raise SystemExit("Nenhum arquivo em data/incoming/")
    cons = {}   # (periodo, code) -> soma
    print("=" * 78)
    print(f"Adaptador DRE — Ref Comunicação — ATUAIS Jan..{MAX_PERIOD[-2:]}/2026 (projeções descartadas)")
    print("=" * 78)
    for f in files:
        facts, ucode, uname, periods = extract_file(f)
        codes = sorted({x["code"] for x in facts})
        tot = {}
        for x in facts:
            tot[x["code"]] = tot.get(x["code"], 0.0) + x["valor"]
            cons[(x["periodo"], x["code"])] = cons.get((x["periodo"], x["code"]), 0.0) + x["valor"]
        miss = [c for c, *_ in LINES if c not in codes]
        print(f"\n■ {uname:5} ({os.path.basename(f)})  meses={len(periods)}  linhas_casadas={len(codes)}/{len(LINES)}")
        print(f"   Receita Bruta {_fmt(tot.get('receita_bruta',0)):>13} | "
              f"Rec.Líq {_fmt(tot.get('receita_liquida',0)):>13} | "
              f"Result.Agência {_fmt(tot.get('resultado_agencia',0)):>13}")
        print(f"   EBIT {_fmt(tot.get('ebit',0)):>13} | "
              f"Result.Líquido {_fmt(tot.get('resultado_liquido',0)):>13} | "
              f"Ger.Caixa {_fmt(tot.get('geracao_caixa',0)):>13}")
        if miss:
            print(f"   ⚠ não casou: {miss}")

    print("\n" + "-" * 78)
    print("CONSOLIDADO (soma das 5 unidades) — totais ATUAIS 2026:")
    ct = {}
    for (per, code), v in cons.items():
        ct[code] = ct.get(code, 0.0) + v
    for code, lbl, *_ in LINES:
        if code in ct:
            print(f"   {lbl:34} {_fmt(ct[code]):>15}")


if __name__ == "__main__":
    import sys
    if "--json" in sys.argv:
        print(json.dumps(build_all_facts(), ensure_ascii=False))
    else:
        main()
